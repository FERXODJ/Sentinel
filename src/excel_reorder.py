from __future__ import annotations

import re
import unicodedata
from pathlib import Path
from typing import Dict, Iterable, List, Tuple

from openpyxl.reader.excel import load_workbook


def _norm_key(value: str) -> str:
    """Normaliza encabezados para hacer matching tolerante.

    - Lower
    - Sin acentos
    - Reemplaza puntuación por espacios
    - Elimina stopwords comunes (de, del, y, etc.)
    """

    s = (value or "").strip().lower()
    if not s:
        return ""

    s = "".join(c for c in unicodedata.normalize("NFKD", s) if not unicodedata.combining(c))
    s = re.sub(r"[^a-z0-9]+", " ", s)
    tokens = [t for t in s.split() if t]

    stop = {
        "de",
        "del",
        "la",
        "el",
        "los",
        "las",
        "y",
        "a",
        "al",
    }
    tokens = [t for t in tokens if t not in stop]
    return " ".join(tokens)


def _sheet_headers(ws) -> List[str]:
    headers: List[str] = []
    for c in range(1, ws.max_column + 1):
        v = ws.cell(row=1, column=c).value
        if v is None:
            headers.append("")
        else:
            headers.append(str(v).strip())

    # recorta trailing vacíos
    while headers and not headers[-1]:
        headers.pop()
    return headers


def _build_col_index(headers: Iterable[str]) -> Dict[str, int]:
    """Devuelve norm_key(header) -> 1-based col index."""
    m: Dict[str, int] = {}
    for idx0, h in enumerate(headers):
        key = _norm_key(h)
        if key and key not in m:
            m[key] = idx0 + 1
    return m


def reorder_datos_completos_by_template(
    *,
    excel_path: str | Path,
    template_path: str | Path,
    datos_completos_sheet: str = "Datos Completos",
    template_sheet: str | None = None,
    keep_extra_columns: bool = True,
    exclude_columns: Iterable[str] | None = None,
) -> Tuple[int, int]:
    """Reordena/renombra columnas de 'Datos Completos' según una plantilla.

    - El orden y nombres de columnas se toman desde la fila 1 del template.
    - Si una columna del template no existe en Datos Completos, se agrega vacía.
    - Opcionalmente conserva columnas extras al final (keep_extra_columns=True).
    - Permite excluir columnas específicas por nombre (exclude_columns).

    Returns: (rows_copied, out_columns)
    """

    excel_path = Path(excel_path)
    template_path = Path(template_path)

    if not excel_path.exists():
        raise FileNotFoundError(f"No existe el archivo: {excel_path}")
    if not template_path.exists():
        raise FileNotFoundError(f"No existe el archivo plantilla: {template_path}")

    wb = load_workbook(excel_path)
    if datos_completos_sheet not in wb.sheetnames:
        raise KeyError(f"No existe la hoja '{datos_completos_sheet}'. Primero ejecuta el merge.")

    ws_src = wb[datos_completos_sheet]

    wb_tpl = load_workbook(template_path, read_only=True, data_only=True)
    if template_sheet:
        if template_sheet not in wb_tpl.sheetnames:
            raise KeyError(f"La plantilla no tiene la hoja '{template_sheet}'.")
        ws_tpl = wb_tpl[template_sheet]
    else:
        ws_tpl = wb_tpl.active

    tpl_headers_raw = [h for h in _sheet_headers(ws_tpl) if h]
    if not tpl_headers_raw:
        raise ValueError("La plantilla no tiene encabezados en la fila 1.")

    excluded_keys: set[str] = set()
    if exclude_columns:
        excluded_keys = {_norm_key(c) for c in exclude_columns if _norm_key(c)}

    if excluded_keys:
        tpl_headers_raw = [h for h in tpl_headers_raw if _norm_key(h) not in excluded_keys]

    src_headers_raw = _sheet_headers(ws_src)
    src_index = _build_col_index(src_headers_raw)

    # Alias manuales: template_key -> source_key
    # (porque hay casos donde el texto cambia mucho: 'fecha y hora de actualización' vs 'Actualizado (fecha y hora)')
    alias: Dict[str, str] = {
        _norm_key("fecha y hora de actualización"): _norm_key("Actualizado (fecha y hora)"),
        _norm_key("actualizacion"): _norm_key("Actualizado (fecha y hora)"),
        _norm_key("creado de fecha y hora"): _norm_key("Creado (fecha y hora)"),
        _norm_key("socio"): _norm_key("Socio"),
        _norm_key("incoming customer"): _norm_key("Incoming Customer"),
        _norm_key("id de cliente"): _norm_key("ID Cliente"),
        _norm_key("categoria del cierre"): _norm_key("Categoria del Cierre"),
        _norm_key("promocion"): _norm_key("Promocion"),
        _norm_key("task"): _norm_key("Task"),
    }

    tpl_norms = [_norm_key(h) for h in tpl_headers_raw]
    tpl_norm_set = {k for k in tpl_norms if k}

    # Todas las columnas del origen que quedan "cubiertas" por la plantilla,
    # incluso si se mapean por alias (para evitar duplicados como Actualizado/actualización).
    covered_source_keys = set(tpl_norm_set)
    for tpl_key in tpl_norms:
        if not tpl_key:
            continue
        covered_source_keys.add(alias.get(tpl_key, tpl_key))

    extras: List[str] = []
    if keep_extra_columns:
        seen_extra_keys: set[str] = set()
        for h in src_headers_raw:
            if not h:
                continue
            src_key = _norm_key(h)
            if excluded_keys and src_key in excluded_keys:
                continue
            if not src_key or src_key in covered_source_keys:
                continue
            if src_key in seen_extra_keys:
                continue
            seen_extra_keys.add(src_key)
            extras.append(h)

    out_headers = [*tpl_headers_raw, *extras]

    # Construye map de salida: out_col_idx -> src_col_idx (1-based) o None
    src_cols_for_out: List[int | None] = []
    for out_h in out_headers:
        out_key = _norm_key(out_h)
        src_key = alias.get(out_key, out_key)
        src_col = src_index.get(src_key)
        src_cols_for_out.append(src_col)

    # Crea hoja nueva y reemplaza
    new_title = f"{datos_completos_sheet} (Ordenado)"
    if new_title in wb.sheetnames:
        wb.remove(wb[new_title])
    ws_new = wb.create_sheet(title=new_title)
    ws_new.append(out_headers)

    rows_copied = 0
    for r in range(2, ws_src.max_row + 1):
        row_out: List = []
        for src_col in src_cols_for_out:
            if src_col is None:
                row_out.append("")
            else:
                row_out.append(ws_src.cell(row=r, column=src_col).value)
        ws_new.append(row_out)
        rows_copied += 1

    # Reemplaza la original
    idx_src = wb.sheetnames.index(datos_completos_sheet)
    wb.remove(ws_src)
    ws_new.title = datos_completos_sheet
    
    # Mantener la pestaña en la misma posición aproximada.
    try:
        current_idx = wb.sheetnames.index(datos_completos_sheet)
        wb.move_sheet(ws_new, offset=idx_src - current_idx)
    except Exception:
        pass
    wb._sheets.remove(ws_new)  # type: ignore[attr-defined]
    wb._sheets.insert(idx_src, ws_new)  # type: ignore[attr-defined]

    wb.save(excel_path)
    return rows_copied, len(out_headers)
