from __future__ import annotations

import os
import re
import json
import shutil
import threading
import time
import zipfile
from datetime import date, datetime
from dataclasses import dataclass
from decimal import Decimal, InvalidOperation
from typing import Callable, Any, Protocol
import unicodedata

from playwright.sync_api import sync_playwright, Page

from openpyxl import Workbook
from openpyxl.reader.excel import load_workbook

from .table_extract import extract_table_to_csv, extract_tickets_to_excel, extract_customers_to_excel
from .excel_merge import merge_tickets_customers


MessageSink = Callable[[str], Any]


class LocatorScope(Protocol):
    def locator(self, selector: str):  # playwright's Locator
        ...


@dataclass(frozen=True)
class SplynxConfig:
    login_url: str
    selectors: dict[str, str]
    tables: dict[str, dict[str, Any]]
    browser: dict[str, Any]


class SplynxSession:
    def __init__(self, config: SplynxConfig, message_sink: MessageSink) -> None:
        self._config = config
        self._message = message_sink

        self._extract_events: dict[str, threading.Event] = {
            "table1": threading.Event(),
            "table2": threading.Event(),
        }
        self._extract_modes: dict[str, str] = {
            "table1": "auto",
            "table2": "auto",
        }
        self._shutdown_event = threading.Event()

        self._enrich_event = threading.Event()
        self._enrich_excel_path: str | None = None
        self._enrich_running = False

        self._collect_dates_event = threading.Event()
        self._collect_ticket_id: str | None = None
        self._collect_dates_excel_path: str | None = None
        self._collect_dates_running = False

        self._page_ready = threading.Event()
        self._page: Page | None = None

    def request_extract(self, table_key: str, mode: str = "auto") -> None:
        if table_key not in self._extract_events:
            self._message(f"Tabla desconocida: {table_key}")
            return
        if not self._page_ready.is_set():
            self._message("Aún no está listo el navegador. Espera unos segundos y vuelve a intentar.")
            return
        self._extract_modes[table_key] = str(mode or "auto").lower()
        self._extract_events[table_key].set()

    def shutdown(self) -> None:
        self._shutdown_event.set()
        for ev in self._extract_events.values():
            ev.set()
        self._enrich_event.set()

    def request_enrich_missing(self, *, excel_path: str) -> None:
        if not self._page_ready.is_set():
            self._message("Aún no está listo el navegador. Espera unos segundos y vuelve a intentar.")
            return
        if self._enrich_running:
            self._message("La búsqueda/enriquecimiento ya está ejecutándose.")
            return
        self._enrich_excel_path = str(excel_path)
        self._enrich_event.set()

    def request_collect_dates_nav(self, *, ticket_id: str) -> None:
        """Navega vía Fast Search al ticket (para luego extraer fechas)."""
        if not self._page_ready.is_set():
            self._message("Aún no está listo el navegador. Espera unos segundos y vuelve a intentar.")
            return

        tid = str(ticket_id or "").strip()
        if not tid:
            self._message("Fechas Esc/Cie: no se recibió Ticket ID.")
            return

        self._collect_ticket_id = tid
        self._collect_dates_excel_path = None
        self._collect_dates_event.set()

    def request_collect_dates_from_excel(self, *, excel_path: str) -> None:
        """Completa fechas en 'Datos Completos' (último Escalamiento O&M y último closed)."""
        if not self._page_ready.is_set():
            self._message("Aún no está listo el navegador. Espera unos segundos y vuelve a intentar.")
            return

        path = str(excel_path or "").strip()
        if not path:
            self._message("Fechas Esc/Cie: no se recibió ruta de Excel.")
            return

        if not os.path.exists(path):
            self._message(f"Fechas Esc/Cie: no existe el archivo: {path}")
            return

        if self._collect_dates_running:
            self._message("Fechas Esc/Cie: ya hay una recolección en progreso.")
            return

        self._collect_ticket_id = None
        self._collect_dates_excel_path = path
        self._collect_dates_event.set()

    def run(self, username: str, password: str) -> None:
        channel = str(self._config.browser.get("channel", "msedge"))
        headless = bool(self._config.browser.get("headless", False))

        with sync_playwright() as p:
            browser = p.chromium.launch(channel=channel, headless=headless)
            context = browser.new_context()
            page = context.new_page()
            self._page = page

            page.set_default_timeout(60_000)

            self._message("Cargando página de login...")
            page.goto(self._config.login_url, wait_until="domcontentloaded")

            user_sel = self._config.selectors.get("username", "#login")
            pass_sel = self._config.selectors.get("password", "#password")

            self._message("Colocando usuario/contraseña en el formulario...")
            page.locator(user_sel).fill(username)
            page.locator(pass_sel).fill(password)

            self._page_ready.set()
            self._message(
                "Listo: completa 2FA y presiona Login MANUALMENTE en Edge. "
                "Luego navega a la pantalla de la tabla y usa los botones de extracción."
            )

            try:
                self._event_loop(page)
            finally:
                try:
                    context.close()
                finally:
                    browser.close()

    def _get_scope(self, page: Page) -> LocatorScope:
        # Splynx suele renderizar vistas dentro de iframes con distintos IDs según la pantalla.
        for frame_id in ("opened-page", "list-page", "opened--view-page"):
            frame_loc = page.locator(f"#{frame_id}")
            if frame_loc.count() > 0:
                try:
                    tag = frame_loc.first.evaluate("el => el.tagName.toLowerCase()")
                except Exception:
                    tag = None

                if tag == "iframe":
                    return page.frame_locator(f"#{frame_id}")

        return page

    def _click_any(self, scope: LocatorScope, selectors: list[str]) -> None:
        last_exc: Exception | None = None
        for sel in selectors:
            try:
                loc = scope.locator(sel).first
                loc.wait_for(state="visible")
                loc.scroll_into_view_if_needed()
                try:
                    loc.click()
                except Exception:
                    # Algunos overlays (Select2) pueden bloquear el click normal.
                    try:
                        loc.click(force=True)
                    except Exception:
                        # Último recurso: ejecutar click vía JS.
                        loc.evaluate("el => el.click()")
                return
            except Exception as exc:
                last_exc = exc
                continue
        if last_exc:
            raise last_exc

    def _fill_any(self, scope: LocatorScope, selectors: list[str], text: str) -> None:
        last_exc: Exception | None = None
        for sel in selectors:
            try:
                loc = scope.locator(sel).first
                loc.wait_for(state="visible")
                loc.scroll_into_view_if_needed()
                loc.fill(text)
                return
            except Exception as exc:
                last_exc = exc
                continue
        if last_exc:
            raise last_exc

    def _press_any(self, scope: LocatorScope, selectors: list[str], key: str) -> None:
        last_exc: Exception | None = None
        for sel in selectors:
            try:
                loc = scope.locator(sel).first
                loc.wait_for(state="visible")
                loc.scroll_into_view_if_needed()
                loc.press(key)
                return
            except Exception as exc:
                last_exc = exc
                continue
        if last_exc:
            raise last_exc

    def _wait_nonempty_any(self, scope: LocatorScope, selectors: list[str], timeout_ms: int) -> None:
        last_exc: Exception | None = None
        for sel in selectors:
            try:
                loc = scope.locator(sel).first
                loc.wait_for(state="visible")
                loc.scroll_into_view_if_needed()

                # Espera hasta que el input tenga un valor no vacío.
                # Importante: hacerlo en el contexto correcto (iframe) vía Locator.
                deadline = time.monotonic() + (timeout_ms / 1000.0)
                while True:
                    if time.monotonic() > deadline:
                        raise TimeoutError(f"Timeout esperando valor no vacío en: {sel}")

                    try:
                        val = loc.evaluate(
                            """el => {
                                if (!el) return '';
                                const v1 = (typeof el.value === 'string') ? el.value : '';
                                const v2 = (typeof el.getAttribute === 'function') ? (el.getAttribute('value') || '') : '';
                                return String(v1 || v2 || '').trim();
                            }"""
                        )
                    except Exception:
                        val = ""

                    if isinstance(val, str) and val.strip():
                        break

                    time.sleep(0.25)

                return
            except Exception as exc:
                last_exc = exc
                continue
        if last_exc:
            raise last_exc

    def _wait_enabled_any(self, scope: LocatorScope, selectors: list[str], timeout_ms: int) -> None:
        last_exc: Exception | None = None
        for sel in selectors:
            try:
                loc = scope.locator(sel).first
                loc.wait_for(state="visible")
                loc.scroll_into_view_if_needed()

                deadline = time.monotonic() + (timeout_ms / 1000.0)
                while True:
                    if time.monotonic() > deadline:
                        raise TimeoutError(f"Timeout esperando botón habilitado: {sel}")

                    try:
                        if loc.is_enabled():
                            break
                    except Exception:
                        pass

                    # Fallback para elementos que no soportan is_enabled()
                    try:
                        ok = loc.evaluate(
                            """el => {
                                if (!el) return false;
                                const disabled = el.disabled === true;
                                const aria = (el.getAttribute && (el.getAttribute('aria-disabled') || '')).toLowerCase();
                                const cls = (el.getAttribute && (el.getAttribute('class') || '')).toLowerCase();
                                return !disabled && aria !== 'true' && !cls.includes('disabled');
                            }"""
                        )
                        if bool(ok):
                            break
                    except Exception:
                        pass

                    time.sleep(0.25)

                return
            except Exception as exc:
                last_exc = exc
                continue
        if last_exc:
            raise last_exc

    def _render_text(self, text: str) -> str:
        # Macros simples para fechas:
        # - {today:%d/%m/%Y}
        # - {month_start:%d/%m/%Y}
        today = date.today()
        month_start = today.replace(day=1)

        rendered = text
        rendered = rendered.replace("{today:%d/%m/%Y}", today.strftime("%d/%m/%Y"))
        rendered = rendered.replace("{month_start:%d/%m/%Y}", month_start.strftime("%d/%m/%Y"))
        return rendered

    def _run_step(self, page: Page, scope: LocatorScope, step: Any) -> None:
        # Backward compatible:
        # - string => click
        # - {"action": "click", "selector": "..."}
        # - {"action": "fill", "selector": "...", "text": "..."}
        if isinstance(step, str):
            candidates = [s.strip() for s in step.split("||") if s.strip()]
            if not candidates:
                return
            try:
                self._click_any(scope, candidates)
                return
            except Exception:
                # Algunos dropdowns de Select2 salen fuera del iframe; probamos en la page raíz.
                if scope is not page:
                    self._click_any(page, candidates)
                else:
                    raise
            return

        if isinstance(step, dict):
            action = str(step.get("action", "click")).lower()
            selector = step.get("selector")
            if not selector:
                return

            if isinstance(selector, str):
                candidates = [s.strip() for s in selector.split("||") if s.strip()]
            else:
                candidates = [str(s).strip() for s in selector if str(s).strip()]

            if action == "fill":
                text = self._render_text(str(step.get("text", "")))
                try:
                    self._fill_any(scope, candidates, text)
                except Exception:
                    if scope is not page:
                        self._fill_any(page, candidates, text)
                    else:
                        raise
                return

            if action == "press":
                key = str(step.get("key", "Enter"))
                try:
                    self._press_any(scope, candidates, key)
                except Exception:
                    if scope is not page:
                        self._press_any(page, candidates, key)
                    else:
                        raise
                return

            if action == "wait_enabled":
                timeout_ms = int(step.get("timeout_ms", 120_000))
                try:
                    self._wait_enabled_any(scope, candidates, timeout_ms=timeout_ms)
                except Exception:
                    if scope is not page:
                        self._wait_enabled_any(page, candidates, timeout_ms=timeout_ms)
                    else:
                        raise
                return

            if action == "wait_nonempty":
                timeout_ms = int(step.get("timeout_ms", 300_000))  # 5 minutos por defecto
                try:
                    self._wait_nonempty_any(scope, candidates, timeout_ms=timeout_ms)
                except Exception:
                    if scope is not page:
                        self._wait_nonempty_any(page, candidates, timeout_ms=timeout_ms)
                    else:
                        raise
                return

            # default: click
            try:
                self._click_any(scope, candidates)
            except Exception:
                if scope is not page:
                    self._click_any(page, candidates)
                else:
                    raise
            return

    def _event_loop(self, page: Page) -> None:
        while not self._shutdown_event.is_set():
            if self._extract_events["table1"].is_set():
                self._extract_events["table1"].clear()
                self._do_extract(page, table_key="table1", mode=self._extract_modes.get("table1", "auto"))

            if self._extract_events["table2"].is_set():
                self._extract_events["table2"].clear()
                self._do_extract(page, table_key="table2", mode=self._extract_modes.get("table2", "auto"))

            if self._enrich_event.is_set():
                self._enrich_event.clear()
                self._do_enrich_missing(page)

            if self._collect_dates_event.is_set():
                self._collect_dates_event.clear()
                excel_path = self._collect_dates_excel_path
                tid = self._collect_ticket_id
                self._collect_dates_excel_path = None
                self._collect_ticket_id = None

                if excel_path:
                    self._do_collect_dates_from_excel(page, excel_path=excel_path)
                else:
                    self._do_collect_dates_nav(page, ticket_id=str(tid or ""))

            page.wait_for_timeout(250)

    def _do_collect_dates_nav(self, page: Page, *, ticket_id: str) -> None:
        tid = str(ticket_id or "").strip()
        if not tid:
            self._message("Fechas Esc/Cie: Ticket ID vacío.")
            return

        try:
            self._message(f"Fechas Esc/Cie: abriendo búsqueda y buscando ticket {tid}...")
            self._fast_search_fill(page, tid)
            page.wait_for_timeout(400)
            ok = self._fast_search_pick_ticket(page, tid)
            if not ok:
                self._message(
                    "Fechas Esc/Cie: no pude seleccionar el ticket en los resultados. "
                    "Confirma que el ID existe y que el panel de búsqueda muestra la opción de ticket."
                )
                return

            wanted_digits = self._id_key(tid)
            scope = self._get_scope(page)
            self._message("Fechas Esc/Cie: ticket seleccionado. Abriendo 'Acciones' y mostrando actividades...")
            self._ensure_ticket_activities_visible(page, scope, ticket_id_digits=wanted_digits, timeout_ms=20_000)
            self._message("Fechas Esc/Cie: actividades visibles. Listo para el próximo paso.")
        except Exception as exc:
            self._message(f"Fechas Esc/Cie: error navegando al ticket: {exc}")

    def _open_actions_dropdown(
        self,
        page: Page,
        root: LocatorScope,
        *,
        ticket_id_digits: str,
        timeout_ms: int,
    ) -> None:
        """Abre el dropdown 'Acciones' (debe funcionar con <button> o <a>)."""
        wanted = self._id_key(ticket_id_digits)
        if not wanted:
            raise RuntimeError("ticket_id inválido")

        sidebar_candidates = [
            f"css=#admin_support_tickets_opened_sticky_sidebar_{wanted}",
            f"css=#admin_support_tickets_closed_sticky_sidebar_{wanted}",
            f"css=div[id$='_sticky_sidebar_{wanted}']",
        ]

        def _get_sidebar(scope: LocatorScope):
            for sel in sidebar_candidates:
                try:
                    loc = scope.locator(sel).first
                    if loc.count() > 0 and loc.is_visible():
                        return loc
                except Exception:
                    continue
            return None

        sidebar = _get_sidebar(root)
        if sidebar is None:
            sidebar = _get_sidebar(page)

        actions_candidates = [
            # Preferir por texto dentro del sidebar específico.
            f"css=#admin_support_tickets_opened_sticky_sidebar_{wanted} :is(a,button):has-text('Acciones')",
            f"css=#admin_support_tickets_closed_sticky_sidebar_{wanted} :is(a,button):has-text('Acciones')",
            f"css=#admin_support_tickets_opened_sticky_sidebar_{wanted} :is(a,button):has-text('Actions')",
            f"css=#admin_support_tickets_closed_sticky_sidebar_{wanted} :is(a,button):has-text('Actions')",

            # Fallback por estructura típica del panel-actions.
            f"css=#admin_support_tickets_opened_sticky_sidebar_{wanted} div.panel-actions.row-gap-4.column-gap-4 > div > :is(a,button)",
            f"css=#admin_support_tickets_closed_sticky_sidebar_{wanted} div.panel-actions.row-gap-4.column-gap-4 > div > :is(a,button)",

            # Fallback por clase/atributo de dropdown, pero sin ser tan amplio.
            f"css=#admin_support_tickets_opened_sticky_sidebar_{wanted} :is(a,button).dropdown-toggle",
            f"css=#admin_support_tickets_closed_sticky_sidebar_{wanted} :is(a,button).dropdown-toggle",
            f"css=#admin_support_tickets_opened_sticky_sidebar_{wanted} :is(a,button)[data-bs-toggle='dropdown']",
            f"css=#admin_support_tickets_closed_sticky_sidebar_{wanted} :is(a,button)[data-bs-toggle='dropdown']",

            # XPaths históricos (último recurso)
            f"xpath=//*[@id='admin_support_tickets_opened_sticky_sidebar_{wanted}']/div/div[2]/div/button",
            f"xpath=//*[@id='admin_support_tickets_closed_sticky_sidebar_{wanted}']/div/div[2]/div/button",
        ]

        last_exc: Exception | None = None
        start = time.monotonic()
        while True:
            if (time.monotonic() - start) * 1000.0 > timeout_ms:
                if last_exc:
                    raise last_exc
                raise RuntimeError("timeout abriendo el dropdown 'Acciones'")

            for sel in actions_candidates:
                try:
                    loc = root.locator(sel).first
                    if loc.count() == 0:
                        continue
                    loc.wait_for(state="visible", timeout=2_000)
                    loc.scroll_into_view_if_needed()
                    try:
                        loc.click()
                    except Exception:
                        try:
                            loc.click(force=True)
                        except Exception:
                            loc.evaluate("el => el.click()")

                    # Verificar que el dropdown realmente abrió (aria-expanded o dropdown-menu.show)
                    opened = False
                    check_until = time.monotonic() + 2.0
                    while time.monotonic() < check_until:
                        try:
                            if (loc.get_attribute("aria-expanded") or "").strip().lower() == "true":
                                opened = True
                                break
                        except Exception:
                            pass

                        try:
                            if sidebar is not None:
                                menu = sidebar.locator("css=.dropdown-menu.show").first
                                if menu.count() > 0 and menu.is_visible():
                                    opened = True
                                    break
                        except Exception:
                            pass

                        time.sleep(0.1)

                    if opened:
                        return
                except Exception as exc:
                    last_exc = exc
                    continue

            time.sleep(0.2)

    def _click_menu_item_any(self, page: Page, root: LocatorScope, selectors: list[str]) -> None:
        """Click robusto para items de menú (asegura click en <a> o <button> si el selector apunta al <li>)."""
        last_exc: Exception | None = None

        def _try(scope: LocatorScope, sel: str) -> bool:
            loc = scope.locator(sel).first
            if loc.count() == 0:
                return False
            loc.wait_for(state="visible", timeout=4_000)
            loc.scroll_into_view_if_needed()

            try:
                tag = (loc.evaluate("el => el.tagName.toLowerCase()") or "").strip().lower()
            except Exception:
                tag = ""

            target = loc
            if tag and tag not in ("a", "button"):
                try:
                    child = loc.locator("css=a,button").first
                    if child.count() > 0:
                        child.wait_for(state="visible", timeout=2_000)
                        child.scroll_into_view_if_needed()
                        target = child
                except Exception:
                    pass

            try:
                target.click()
            except Exception:
                try:
                    target.click(force=True)
                except Exception:
                    target.evaluate("el => el.click()")
            return True

        for sel in selectors:
            try:
                if _try(root, sel):
                    return
            except Exception as exc:
                last_exc = exc

            try:
                if _try(page, sel):
                    return
            except Exception as exc:
                last_exc = exc

        if last_exc:
            raise last_exc
        raise RuntimeError("no pude clicar el item del menú")

    def _do_collect_dates_from_excel(self, page: Page, *, excel_path: str) -> None:
        if self._collect_dates_running:
            self._message("Fechas Esc/Cie: ya hay una recolección en progreso.")
            return

        self._collect_dates_running = True
        try:
            try:
                wb = self._open_or_create_workbook(excel_path)
            except PermissionError:
                self._message(
                    "Fechas Esc/Cie: no puedo abrir/guardar el Excel porque está abierto o bloqueado. "
                    "Cierra 'output/Datos Splynx.xlsx' y reintenta."
                )
                return
            except Exception as exc:
                self._message(f"Fechas Esc/Cie: el Excel no es válido o está dañado: {exc}")
                self._message(
                    "Solución: abre el archivo en Excel y usa 'Guardar como' .xlsx, o vuelve a generar el archivo con Extraer/Comparar."
                )
                return
            if "Datos Completos" not in wb.sheetnames:
                self._message(
                    "Fechas Esc/Cie: no existe la hoja 'Datos Completos'. Ejecuta primero 'Comparar y agrupar datos'."
                )
                return

            ws = wb["Datos Completos"]
            if ws.max_row < 2:
                self._message("Fechas Esc/Cie: la hoja 'Datos Completos' está vacía.")
                return

            headers = [str(ws.cell(row=1, column=c).value or "").strip() for c in range(1, ws.max_column + 1)]

            def _find_header(name: str) -> int | None:
                n = (name or "").strip().lower()
                for idx0, h in enumerate(headers):
                    if (h or "").strip().lower() == n:
                        return idx0 + 1
                return None

            id_col = _find_header("ID")
            if not id_col:
                self._message("Fechas Esc/Cie: no encontré la columna 'ID' en 'Datos Completos'.")
                return

            # Asegurar columnas en orden: Escalamiento, Resuelto, Cierre.
            esc_col = _find_header("Fecha Escalamiento (O&M)")
            res_col = _find_header("Resuelto")
            cie_col = _find_header("Fecha Cierre (closed)")

            if not esc_col:
                esc_col = ws.max_column + 1
                ws.cell(row=1, column=esc_col).value = "Fecha Escalamiento (O&M)"
                headers.append("Fecha Escalamiento (O&M)")

            # Resuelto debe ir entre Escalamiento y Cierre.
            if not res_col:
                if cie_col:
                    # Insertar justo antes de Cierre para quedar en el medio.
                    ws.insert_cols(cie_col)
                    res_col = cie_col
                    ws.cell(row=1, column=res_col).value = "Resuelto"
                    headers.insert(res_col - 1, "Resuelto")
                    cie_col = cie_col + 1
                else:
                    res_col = ws.max_column + 1
                    ws.cell(row=1, column=res_col).value = "Resuelto"
                    headers.append("Resuelto")

            if not cie_col:
                cie_col = ws.max_column + 1
                ws.cell(row=1, column=cie_col).value = "Fecha Cierre (closed)"
                headers.append("Fecha Cierre (closed)")

            ticket_rows: list[tuple[int, str]] = []
            for r in range(2, ws.max_row + 1):
                raw = ws.cell(row=r, column=id_col).value
                tid = self._id_key(raw)
                if tid:
                    ticket_rows.append((r, tid))

            if not ticket_rows:
                self._message("Fechas Esc/Cie: no encontré IDs válidos en 'Datos Completos'.")
                return

            total = len(ticket_rows)
            self._message(f"Fechas Esc/Cie: iniciando recolección para {total} tickets...")

            # Backup best-effort para evitar pérdida si hay un corte durante un save.
            self._ensure_excel_backup(excel_path)

            # Reanudar desde checkpoint si existe.
            progress_path = self._progress_path_for_excel(excel_path)
            progress = self._load_progress(progress_path)
            start_row_idx = int(progress.get("last_row_idx") or 0)
            start_index = 0
            if start_row_idx:
                for idx0, (ridx, _) in enumerate(ticket_rows):
                    if ridx == start_row_idx:
                        start_index = min(idx0 + 1, len(ticket_rows))
                        break
                if start_index:
                    self._message(
                        f"Fechas Esc/Cie: reanudando desde fila {start_row_idx} (posición {start_index}/{total})."
                    )

            updated = 0
            skipped = 0
            failed = 0
            saved_at = time.monotonic()

            def _is_missing(v: str) -> bool:
                t = str(v or "").strip()
                if not t:
                    return True
                return t.strip().lower() in ("n/a", "na")

            for i, (row_idx, ticket_id) in enumerate(ticket_rows[start_index:], start=start_index + 1):
                existing_esc = str(ws.cell(row=row_idx, column=esc_col).value or "").strip()
                existing_res = str(ws.cell(row=row_idx, column=res_col).value or "").strip()
                existing_cie = str(ws.cell(row=row_idx, column=cie_col).value or "").strip()

                # Si ya hay valores reales (no vacíos y no N/A) en las 3, saltar.
                if not _is_missing(existing_esc) and not _is_missing(existing_res) and not _is_missing(existing_cie):
                    skipped += 1
                    continue

                self._message(f"Fechas Esc/Cie: [{i}/{total}] Ticket {ticket_id}...")

                esc_dt = ""
                res_dt = ""
                close_dt = ""
                last_exc: Exception | None = None

                for attempt in range(1, 4):
                    try:
                        esc_dt, res_dt, close_dt = self._collect_dates_for_ticket(page, ticket_id=ticket_id)
                        last_exc = None
                        break
                    except Exception as exc:
                        last_exc = exc
                        msg = str(exc).lower()
                        retryable = any(
                            k in msg
                            for k in (
                                "timeout",
                                "fast search",
                                "internet_disconnected",
                                "err_internet_disconnected",
                                "err_network_changed",
                                "err_connection",
                                "net::",
                                "navigation",
                            )
                        )
                        if attempt >= 3 or not retryable:
                            break

                        self._message(
                            f"Fechas Esc/Cie: Ticket {ticket_id}: reintento {attempt}/3 por error: {exc}"
                        )
                        self._recover_page_after_error(page)
                        time.sleep(2.0)

                if last_exc is None:
                    # Si el dato no existe, escribir N/A (para que no se vea vacío)
                    esc_out = esc_dt or "N/A"
                    res_out = res_dt or "N/A"
                    close_out = close_dt or "N/A"

                    if _is_missing(existing_esc):
                        ws.cell(row=row_idx, column=esc_col).value = esc_out
                    if _is_missing(existing_res):
                        ws.cell(row=row_idx, column=res_col).value = res_out
                    if _is_missing(existing_cie):
                        ws.cell(row=row_idx, column=cie_col).value = close_out

                    updated += 1
                    self._message(
                        f"Fechas Esc/Cie: Ticket {ticket_id}: Esc(O&M)={'OK' if esc_dt else 'NO'}, Resuelto={'OK' if res_dt else 'NO'}, Cierre(closed)={'OK' if close_dt else 'NO'}."
                    )
                else:
                    failed += 1
                    self._message(f"Fechas Esc/Cie: Ticket {ticket_id}: error: {last_exc}")

                # Guardar checkpoint (aunque falle) para poder reanudar cerca de donde quedó.
                self._save_progress(
                    progress_path,
                    {
                        "excel": excel_path,
                        "last_row_idx": row_idx,
                        "last_ticket_id": ticket_id,
                        "updated": updated,
                        "skipped": skipped,
                        "failed": failed,
                        "ts": datetime.now().isoformat(timespec="seconds"),
                    },
                )

                if (updated + failed) % 10 == 0 or (time.monotonic() - saved_at) > 30:
                    try:
                        self._atomic_save_workbook(wb, excel_path)
                        saved_at = time.monotonic()
                    except PermissionError:
                        self._message(
                            "Fechas Esc/Cie: no pude guardar el Excel (está abierto/bloqueado). Cierra el archivo y reintenta."
                        )
                        return

            try:
                self._atomic_save_workbook(wb, excel_path)
            except PermissionError:
                self._message(
                    "Fechas Esc/Cie: no pude guardar el Excel al final (está abierto/bloqueado). Cierra el archivo y reintenta."
                )
                return

            # Marcar checkpoint final como terminado.
            self._save_progress(
                progress_path,
                {
                    "excel": excel_path,
                    "done": True,
                    "updated": updated,
                    "skipped": skipped,
                    "failed": failed,
                    "ts": datetime.now().isoformat(timespec="seconds"),
                },
            )

            self._message(f"Fechas Esc/Cie: terminado. Actualizados: {updated}, saltados: {skipped}, fallos: {failed}.")
        finally:
            self._collect_dates_running = False

    def _recover_page_after_error(self, page: Page) -> None:
        """Recovery best-effort cuando la UI se queda pegada o hay cortes de internet."""
        try:
            page.keyboard.press("Escape")
            page.wait_for_timeout(150)
            page.keyboard.press("Escape")
            page.wait_for_timeout(150)
        except Exception:
            pass

        # Intentar recargar sin bloquear demasiado.
        try:
            page.reload(wait_until="domcontentloaded")
            page.wait_for_timeout(400)
        except Exception:
            pass

    def _progress_path_for_excel(self, excel_path: str) -> str:
        # Guardar al lado del Excel para que sea fácil ubicarlo.
        return f"{excel_path}.progress.json"

    def _load_progress(self, path: str) -> dict[str, object]:
        try:
            if not os.path.exists(path):
                return {}
            with open(path, "r", encoding="utf-8") as f:
                data = json.load(f)
            return data if isinstance(data, dict) else {}
        except Exception:
            return {}

    def _save_progress(self, path: str, data: dict[str, object]) -> None:
        try:
            tmp = f"{path}.tmp"
            os.makedirs(os.path.dirname(path) or ".", exist_ok=True)
            with open(tmp, "w", encoding="utf-8") as f:
                json.dump(data, f, ensure_ascii=False, indent=2)
            os.replace(tmp, path)
        except Exception:
            pass

    def _ensure_excel_backup(self, excel_path: str) -> None:
        """Crea un .bak una sola vez (best-effort)."""
        try:
            if not os.path.exists(excel_path):
                return
            bak = f"{excel_path}.bak"
            if os.path.exists(bak):
                return
            shutil.copy2(excel_path, bak)
            self._message(f"Fechas Esc/Cie: backup creado: {bak}")
        except Exception:
            # No bloquear la ejecución si no se puede respaldar.
            pass

    def _atomic_save_workbook(self, wb: Workbook, excel_path: str) -> None:
        """Guarda el .xlsx de forma atómica (temp + replace) para evitar archivos corruptos."""
        tmp = f"{excel_path}.tmp"
        wb.save(tmp)
        os.replace(tmp, excel_path)

    def _collect_dates_for_ticket(self, page: Page, *, ticket_id: str) -> tuple[str, str, str]:
        tid = str(ticket_id or "").strip()
        if not tid:
            return ("", "")

        self._fast_search_fill(page, tid)
        page.wait_for_timeout(400)
        if not self._fast_search_pick_ticket(page, tid):
            raise RuntimeError("no pude seleccionar el ticket en el Fast Search")

        # Dar tiempo a que Splynx navegue/renderice la vista seleccionada.
        page.wait_for_timeout(250)

        wanted_digits = self._id_key(tid)
        if wanted_digits:
            if not self._wait_ticket_view_loaded_for_id(page, wanted_digits, timeout_ms=35_000):
                raise RuntimeError("timeout esperando que cargue el ticket correcto")

        # A veces la búsqueda abre un modal ("Búsqueda") que queda encima y bloquea clicks.
        # Cerrarlo best-effort antes de interactuar con el panel lateral.
        try:
            page.keyboard.press("Escape")
            page.wait_for_timeout(120)
            page.keyboard.press("Escape")
            page.wait_for_timeout(120)
        except Exception:
            pass

        # El scope puede cambiar al navegar; re-evalúalo antes de interactuar.
        scope = self._get_scope(page)
        self._ensure_ticket_activities_visible(page, scope, ticket_id_digits=wanted_digits, timeout_ms=30_000)

        # Verificación extra: esperar a que activities realmente cargue contenido.
        # Esto evita continuar “demasiado rápido” cuando el dropdown se clickea pero el historial aún no renderiza.
        blocks = self._wait_activities_loaded(page, scope, timeout_ms=25_000, min_blocks=1)
        if blocks <= 0:
            raise RuntimeError("no se cargaron actividades (historial vacío/no visible)")

        # Mensaje breve de certificación (ayuda a auditar que sí hubo carga de historial).
        try:
            self._message(f"Fechas Esc/Cie: Ticket {tid}: activities cargadas ({blocks} bloques).")
        except Exception:
            pass

        esc_dt = self._extract_last_escalation_om(scope, page)
        res_dt = self._extract_last_resuelto(scope, page)
        close_dt = self._extract_last_closed(scope, page)

        # Algunos tickets (especialmente cerrados) cargan el historial de forma parcial/lazy.
        # Si no aparece el evento de cierre, intentar cargar más activities (scroll) y reintentar 1 vez.
        if not close_dt or not res_dt:
            try:
                self._message(f"Fechas Esc/Cie: Ticket {tid}: cargando más activities para validar cierre/resuelto...")
            except Exception:
                pass
            self._load_more_activities_by_scrolling(page)
            if not res_dt:
                res_dt = self._extract_last_resuelto(scope, page)
            close_dt = self._extract_last_closed(scope, page)
        return (esc_dt, res_dt, close_dt)

    def _load_more_activities_by_scrolling(self, page: Page) -> None:
        """Best-effort: hace scroll para disparar carga lazy de historial."""
        try:
            # Hacer varios scrolls hacia abajo suele cargar más bloques.
            for _ in range(4):
                try:
                    page.mouse.wheel(0, 2200)
                except Exception:
                    try:
                        page.evaluate("() => window.scrollBy(0, 2200)")
                    except Exception:
                        pass
                page.wait_for_timeout(250)

            # Y un poco hacia arriba para estabilizar.
            for _ in range(2):
                try:
                    page.mouse.wheel(0, -1200)
                except Exception:
                    try:
                        page.evaluate("() => window.scrollBy(0, -1200)")
                    except Exception:
                        pass
                page.wait_for_timeout(200)
        except Exception:
            pass

    def _has_word(self, norm: str, word: str) -> bool:
        w = (word or "").strip().lower()
        if not w:
            return False
        # límites alfanuméricos (evita coincidir dentro de otras palabras)
        return bool(re.search(rf"(?<![a-z0-9]){re.escape(w)}(?![a-z0-9])", norm or ""))

    def _count_activity_blocks(self, page: Page, scope: LocatorScope) -> int:
        sel = "css=div[id^='opened-ticket-message-']"
        best = 0
        try:
            c = scope.locator(sel).count()
            best = max(best, int(c or 0))
        except Exception:
            pass
        try:
            c = page.locator(sel).count()
            best = max(best, int(c or 0))
        except Exception:
            pass
        return best

    def _wait_activities_loaded(self, page: Page, scope: LocatorScope, *, timeout_ms: int, min_blocks: int = 1) -> int:
        """Espera a que el historial/activities tenga bloques renderizados.

        Retorna el número de bloques detectados (máximo entre iframe y page).
        """
        start = time.monotonic()
        last = 0
        while True:
            if (time.monotonic() - start) * 1000.0 > timeout_ms:
                return last
            last = self._count_activity_blocks(page, scope)
            if last >= int(min_blocks or 1):
                return last
            time.sleep(0.25)

    def _wait_ticket_view_loaded_for_id(self, page: Page, ticket_id_digits: str, timeout_ms: int) -> bool:
        wanted = self._id_key(ticket_id_digits)
        if not wanted:
            return True

        # En algunas vistas, Splynx renderiza dentro de iframes y la URL del page
        # no siempre refleja el ticket actual. Mejor esperar por elementos del DOM del ticket.
        selectors = [
            f"css=#admin_support_tickets_opened_sticky_sidebar_{wanted}",
            f"css=#admin_support_tickets_closed_sticky_sidebar_{wanted}",
            f"css=#admin_support_tickets_opened_view_show_hide_activities_{wanted}",
            f"css=#admin_support_tickets_closed_view_show_hide_activities_{wanted}",
        ]
        start = time.monotonic()
        while True:
            if (time.monotonic() - start) * 1000.0 > timeout_ms:
                return False

            # Señal 1: URL con ?id=<wanted> (cuando aplica)
            try:
                u = page.url or ""
                if re.search(rf"[?&]id={re.escape(wanted)}(?:$|[^0-9])", u):
                    return True
            except Exception:
                pass

            # Señal 2: DOM del ticket
            if self._wait_visible_any(page, selectors, timeout_ms=1_000):
                return True

            time.sleep(0.2)

    def _ensure_ticket_activities_visible(
        self,
        page: Page,
        scope: LocatorScope,
        *,
        ticket_id_digits: str,
        timeout_ms: int,
    ) -> None:
        wanted = self._id_key(ticket_id_digits)
        if not wanted:
            raise RuntimeError("ticket_id inválido")

        # 1) Abrir el dropdown de 'Acciones' (verificando que efectivamente abre)
        try:
            self._open_actions_dropdown(page, scope, ticket_id_digits=wanted, timeout_ms=min(timeout_ms, 12_000))
        except Exception:
            self._open_actions_dropdown(page, page, ticket_id_digits=wanted, timeout_ms=min(timeout_ms, 12_000))
        page.wait_for_timeout(200)

        # 2) Click en Show activities
        # Nota: a veces el ID está en el <li> y el clickable real es el <a> dentro.
        activities_candidates = [
            # Preferir <a> dentro del item con ID
            f"css=#admin_support_tickets_opened_view_show_hide_activities_{wanted} a",
            f"css=#admin_support_tickets_closed_view_show_hide_activities_{wanted} a",
            f"css=#admin_support_tickets_opened_view_show_hide_activities_{wanted} :is(a,button)",
            f"css=#admin_support_tickets_closed_view_show_hide_activities_{wanted} :is(a,button)",

            # Fallback al nodo con ID
            f"css=#admin_support_tickets_opened_view_show_hide_activities_{wanted}",
            f"css=#admin_support_tickets_closed_view_show_hide_activities_{wanted}",
            f"css=[id$='_view_show_hide_activities_{wanted}'] a",
            f"css=[id$='_view_show_hide_activities_{wanted}'] :is(a,button)",
            f"css=[id$='_view_show_hide_activities_{wanted}']",

            # XPaths por ID
            f"xpath=//*[@id='admin_support_tickets_opened_view_show_hide_activities_{wanted}']/a",
            f"xpath=//*[@id='admin_support_tickets_closed_view_show_hide_activities_{wanted}']/a",
            f"xpath=//*[@id='admin_support_tickets_opened_view_show_hide_activities_{wanted}']",
            f"xpath=//*[@id='admin_support_tickets_closed_view_show_hide_activities_{wanted}']",

            # Último recurso: XPath absoluto provisto (puede variar según layout)
            "xpath=/html/body/div[2]/div[3]/div[1]/div/div/div[2]/div[2]/div[1]/div[1]/div/div[2]/div/ul/li[1]/a",
        ]

        # La opción está dentro del dropdown; esperamos a que sea visible antes del click.
        if not self._wait_visible_any(page, activities_candidates, timeout_ms=min(timeout_ms, 12_000)):
            # Reintento: abrir el menú otra vez (a veces el primer click no abre por overlay)
            try:
                self._open_actions_dropdown(page, scope, ticket_id_digits=wanted, timeout_ms=6_000)
            except Exception:
                self._open_actions_dropdown(page, page, ticket_id_digits=wanted, timeout_ms=6_000)
            page.wait_for_timeout(200)

        if not self._wait_visible_any(page, activities_candidates, timeout_ms=min(timeout_ms, 12_000)):
            raise RuntimeError("no pude encontrar la opción 'Show activities' en el menú Acciones")

        # Click robusto (si el selector apunta a <li>, hace click al <a> interno)
        self._click_menu_item_any(page, scope, activities_candidates)

        # Post-check: esperar a que el dropdown se cierre (si no, Escape best-effort)
        try:
            page.wait_for_timeout(150)
            sidebar_sel = f"css=#admin_support_tickets_opened_sticky_sidebar_{wanted}, #admin_support_tickets_closed_sticky_sidebar_{wanted}, div[id$='_sticky_sidebar_{wanted}']"
            sidebar = page.locator(sidebar_sel).first
            if sidebar.count() > 0:
                start = time.monotonic()
                while (time.monotonic() - start) < 2.0:
                    try:
                        menu = sidebar.locator("css=.dropdown-menu.show").first
                        if menu.count() == 0 or not menu.is_visible():
                            break
                    except Exception:
                        break
                    time.sleep(0.1)
        except Exception:
            pass

        try:
            page.keyboard.press("Escape")
        except Exception:
            pass

        # Darle tiempo a Splynx a renderizar el historial.
        page.wait_for_timeout(600)

        # Espera/verificación: confirmar que aparezcan bloques de historial.
        # Si no aparecen, reintentar el click de Show activities una vez.
        blocks = self._wait_activities_loaded(page, scope, timeout_ms=min(timeout_ms, 15_000), min_blocks=1)
        if blocks <= 0:
            try:
                self._open_actions_dropdown(page, scope, ticket_id_digits=wanted, timeout_ms=6_000)
            except Exception:
                self._open_actions_dropdown(page, page, ticket_id_digits=wanted, timeout_ms=6_000)
            page.wait_for_timeout(150)
            self._click_menu_item_any(page, scope, activities_candidates)
            page.wait_for_timeout(600)
            blocks = self._wait_activities_loaded(page, scope, timeout_ms=min(timeout_ms, 15_000), min_blocks=1)

        if blocks <= 0:
            raise RuntimeError("no se pudo cargar el historial de activities tras 'Show activities'")

    def _extract_last_escalation_om(self, scope: LocatorScope, page: Page) -> str:
        return self._extract_last_activity_datetime_for_match(
            scope,
            page,
            match_predicate=lambda norm: ("changed group" in norm or "cambiado grupo" in norm or "cambio grupo" in norm)
            and ("operacion y mantenimiento" in norm or "operación y mantenimiento" in norm),
        )

    def _extract_last_resuelto(self, scope: LocatorScope, page: Page) -> str:
        def _pred(norm: str) -> bool:
            n = norm or ""
            # Estado destino: Resuelto como palabra completa
            if not self._has_word(n, "resuelto"):
                return False

            # Contexto: preferir eventos de cambio de estado.
            # Si la UI cambia el texto, aceptar también si menciona "status" o "estado".
            has_context = any(
                k in n
                for k in (
                    "changed status",
                    "status changed",
                    "cambiado estado",
                    "cambio estado",
                    "cambiar el estado",
                )
            )
            return has_context or ("status" in n) or ("estado" in n)

        return self._extract_last_activity_datetime_for_match(scope, page, match_predicate=_pred)

    def _extract_last_closed(self, scope: LocatorScope, page: Page) -> str:
        def _pred(norm: str) -> bool:
            n = norm or ""

            # Contextos típicos de cambio/cierre
            has_context = any(
                k in n
                for k in (
                    "changed status",
                    "status changed",
                    "ticket closed",
                    "closed ticket",
                    "cambiado estado",
                    "cambio estado",
                    "cambiar el estado",
                    "ticket cerrado",
                    "cerrado el ticket",
                )
            )
            if not has_context:
                return False

            # Estado destino: closed/cerrado como palabra completa
            return self._has_word(n, "closed") or self._has_word(n, "cerrado")

        return self._extract_last_activity_datetime_for_match(scope, page, match_predicate=_pred)

    def _extract_last_activity_datetime_for_match(
        self,
        scope: LocatorScope,
        page: Page,
        *,
        match_predicate,
    ) -> str:
        blocks_selector = "css=div[id^='opened-ticket-message-']"
        dt_selector = "css=div.comment-heading div.comment-title-wrapper span"
        dt_selector_fallbacks = [
            dt_selector,
            "css=div.comment-heading span",
        ]

        def _get_blocks(container: LocatorScope):
            return container.locator(blocks_selector)

        blocks = _get_blocks(scope)
        try:
            if blocks.count() == 0:
                blocks = _get_blocks(page)
        except Exception:
            blocks = _get_blocks(page)

        best_dt: datetime | None = None
        best_str = ""

        try:
            count = blocks.count()
        except Exception:
            count = 0

        for i in range(min(count, 400)):
            blk = blocks.nth(i)
            try:
                # text_content permite leer aunque el bloque esté oculto (activities colapsadas)
                txt = blk.text_content() or ""
            except Exception:
                continue

            norm = self._norm_text(txt)
            if not match_predicate(norm):
                continue

            try:
                dt_raw = ""
                for dsel in dt_selector_fallbacks:
                    try:
                        dt_raw = (blk.locator(dsel).first.text_content() or "").strip()
                    except Exception:
                        dt_raw = ""
                    if dt_raw:
                        break
            except Exception:
                dt_raw = ""

            dt = self._parse_activity_datetime(dt_raw)
            if not dt:
                # Fallback: en algunos casos el timestamp no está en el heading como esperamos,
                # pero sí aparece dentro del texto completo del bloque.
                dt = self._parse_activity_datetime(txt)
            if not dt:
                continue

            if best_dt is None or dt > best_dt:
                best_dt = dt
                best_str = dt.strftime("%d/%m/%Y %H:%M")

        return best_str

    def _parse_activity_datetime(self, s: str) -> datetime | None:
        text = (s or "").strip()
        if not text:
            return None

        candidates = re.findall(r"\((\d{2}/\d{2}/\d{4}\s+\d{1,2}:\d{2}:\d{2}\s*(?:AM|PM)?)\)", text)
        if not candidates:
            candidates = re.findall(r"(\d{2}/\d{2}/\d{4}\s+\d{1,2}:\d{2}:\d{2}\s*(?:AM|PM)?)", text)

        for cand in reversed(candidates):
            c = cand.strip()
            if not c:
                continue
            for fmt in (
                "%d/%m/%Y %I:%M:%S %p",
                "%d/%m/%Y %H:%M:%S",
                "%d/%m/%Y %I:%M:%S%p",
                "%d/%m/%Y %H:%M",
                "%d/%m/%Y %I:%M %p",
            ):
                try:
                    return datetime.strptime(c, fmt)
                except Exception:
                    pass

        return None

    def _norm_text(self, s: str) -> str:
        s = (s or "").strip().lower()
        try:
            s = "".join(c for c in unicodedata.normalize("NFKD", s) if not unicodedata.combining(c))
        except Exception:
            pass
        s = re.sub(r"\s+", " ", s)
        return s

    _SCI_RE = re.compile(r"^\s*[-+]?\d+(?:\.\d+)?[eE][-+]?\d+\s*$")

    def _id_key(self, value) -> str:
        if value is None:
            return ""
        if isinstance(value, bool):
            return ""
        if isinstance(value, int):
            return str(value)
        if isinstance(value, float):
            try:
                return str(int(value))
            except Exception:
                pass

        s = str(value).strip()
        if not s:
            return ""

        if self._SCI_RE.match(s):
            try:
                num = Decimal(s)
                s = str(int(num))
            except (InvalidOperation, ValueError):
                pass

        matches = re.findall(r"\d+", s)
        if not matches:
            return ""
        d = max(matches, key=len)
        d2 = d.lstrip("0")
        return d2 if d2 else "0"

    def _open_or_create_workbook(self, path: str) -> Workbook:
        if not os.path.exists(path):
            return Workbook()
        try:
            return load_workbook(path)
        except zipfile.BadZipFile as exc:
            raise ValueError(
                f"El archivo '{path}' no es un .xlsx válido (parece estar corrupto o no es un archivo de Excel)."
            ) from exc
        except PermissionError as exc:
            raise PermissionError(
                f"No se puede abrir '{path}'. Probablemente está abierto en Excel o bloqueado. "
                "Ciérralo y vuelve a intentar."
            ) from exc

    def _get_or_create_sheet(self, wb: Workbook, sheet_name: str):
        if sheet_name in wb.sheetnames:
            return wb[sheet_name]

        # Si el workbook recién creado solo tiene Sheet vacío, reutiliza
        if wb.sheetnames == ["Sheet"] and wb["Sheet"].max_row == 1 and wb["Sheet"].max_column == 1:
            ws = wb["Sheet"]
            ws.title = sheet_name
            return ws
        return wb.create_sheet(title=sheet_name)

    def _read_missing_ids(self, excel_path: str) -> list[str]:
        wb = load_workbook(excel_path)
        if "Datos no Encontrados" not in wb.sheetnames:
            raise KeyError("No existe la hoja 'Datos no Encontrados'. Ejecuta primero 'Comparar y agrupar datos'.")

        ws = wb["Datos no Encontrados"]
        headers = [str(ws.cell(row=1, column=c).value or "").strip() for c in range(1, ws.max_column + 1)]
        norm_to_col: dict[str, int] = {}
        for idx0, h in enumerate(headers):
            key = self._norm_text(h)
            if key and key not in norm_to_col:
                norm_to_col[key] = idx0 + 1

        def _col(name: str) -> int | None:
            return norm_to_col.get(self._norm_text(name))

        c_reporter_id = _col("Reporter ID")
        c_reporter_type = _col("Reporter type")
        c_id_cliente = _col("ID Cliente")

        if not c_reporter_id and not c_id_cliente:
            raise KeyError("En 'Datos no Encontrados' no existe 'Reporter ID' ni 'ID Cliente'.")

        ids: list[str] = []
        seen: set[str] = set()

        for r in range(2, ws.max_row + 1):
            reporter_id_val = self._id_key(ws.cell(row=r, column=c_reporter_id).value) if c_reporter_id else ""
            reporter_type_val = (
                self._norm_text(ws.cell(row=r, column=c_reporter_type).value)
                if c_reporter_type
                else ""
            )
            id_cliente_val = self._id_key(ws.cell(row=r, column=c_id_cliente).value) if c_id_cliente else ""

            join_id = ""
            if reporter_type_val == "customer" and reporter_id_val:
                join_id = reporter_id_val
            elif id_cliente_val:
                join_id = id_cliente_val
            elif reporter_id_val:
                join_id = reporter_id_val

            if join_id and join_id not in seen:
                seen.add(join_id)
                ids.append(join_id)

        return ids

    def _fast_search_open(self, page: Page) -> None:
        # Nota: este botón suele ser toggle. Por eso NO debemos clickear si el panel ya está abierto.
        if self._fast_search_is_open(page):
            return

        selectors = [
            "css=body > div.splynx-wrapper > div.splynx-header > ul > li:nth-child(2)",
            # fallbacks menos frágiles (ícono de búsqueda en header)
            "css=div.splynx-header ul li:has(i.fa-search)",
            "css=div.splynx-header ul li:has(i.fa.fa-search)",
            "css=div.splynx-header :is(a,button):has(i.fa-search)",
            "css=div.splynx-header :is(a,button):has-text('Search')",
            "xpath=//*[@id='dashboard-page']/body/div[2]/div[2]/ul/li[2]",
            "xpath=/html/body/div[2]/div[2]/ul/li[2]",
            "xpath=//*[@id='opened--view-page']/body/div[2]/div[2]/ul/li[2]",
        ]
        self._click_any(page, selectors)

    def _wait_fast_search_input_visible(self, page: Page, timeout_ms: int) -> bool:
        selectors = [
            "css=body > div.splynx-wrapper > div.sidebar-wrapper > div > div.sidebar-content > div > div.search-wrapper > div > input",
            "css=div.search-wrapper input",
        ]
        start = time.monotonic()
        while True:
            if (time.monotonic() - start) * 1000.0 > timeout_ms:
                return False
            for sel in selectors:
                try:
                    loc = page.locator(sel).first
                    if loc.count() > 0 and loc.is_visible():
                        return True
                except Exception:
                    continue
            time.sleep(0.2)

    def _fast_search_is_open(self, page: Page) -> bool:
        selectors = [
            "css=body > div.splynx-wrapper > div.sidebar-wrapper > div > div.sidebar-content > div > div.search-wrapper > div > input",
            "css=div.search-wrapper input",
        ]
        for sel in selectors:
            try:
                loc = page.locator(sel).first
                if loc.count() > 0 and loc.is_visible():
                    return True
            except Exception:
                continue
        return False

    def _fast_search_clear(self, page: Page) -> None:
        selectors = [
            "css=body > div.splynx-wrapper > div.sidebar-wrapper > div > div.sidebar-content > div > div.search-wrapper > div > input",
            "xpath=//*[@id='dashboard-page']/body/div[2]/div[5]/div/div[2]/div/div[1]/div/input",
            "xpath=/html/body/div[2]/div[5]/div/div[2]/div/div[1]/div/input",
            "xpath=//*[@id='opened--view-page']/body/div[2]/div[5]/div/div[2]/div/div[1]/div/input",
            "css=div.search-wrapper input",
        ]
        try:
            self._click_any(page, selectors)
            self._fill_any(page, selectors, "")
        except Exception:
            # best-effort
            pass

    def _fast_search_fill(self, page: Page, customer_id: str) -> None:
        # Asegurar panel abierto y limpiar búsqueda previa.
        # Si el input no aparece (por overlays o estado raro del UI), reintentar y auto-recuperar.
        def _ensure_open() -> None:
            if self._fast_search_is_open(page):
                return
            self._fast_search_open(page)

        # intento 1
        _ensure_open()
        if not self._wait_fast_search_input_visible(page, timeout_ms=6_000):
            # cerrar overlays y reintentar
            try:
                page.keyboard.press("Escape")
                page.wait_for_timeout(120)
                page.keyboard.press("Escape")
                page.wait_for_timeout(120)
            except Exception:
                pass

            _ensure_open()

        if not self._wait_fast_search_input_visible(page, timeout_ms=6_000):
            # último recurso: recargar y volver a intentar (a veces el sidebar se rompe tras muchos tickets)
            try:
                page.reload(wait_until="domcontentloaded")
                page.wait_for_timeout(400)
            except Exception:
                pass
            _ensure_open()

        if not self._wait_fast_search_input_visible(page, timeout_ms=10_000):
            raise RuntimeError("Fast Search no está visible (no encuentro el input de búsqueda)")

        page.wait_for_timeout(150)
        self._fast_search_clear(page)

        selectors = [
            "css=body > div.splynx-wrapper > div.sidebar-wrapper > div > div.sidebar-content > div > div.search-wrapper > div > input",
            "xpath=//*[@id='dashboard-page']/body/div[2]/div[5]/div/div[2]/div/div[1]/div/input",
            "xpath=/html/body/div[2]/div[5]/div/div[2]/div/div[1]/div/input",
            "xpath=//*[@id='opened--view-page']/body/div[2]/div[5]/div/div[2]/div/div[1]/div/input",
            "css=div.search-wrapper input",
        ]
        self._click_any(page, selectors)
        self._fill_any(page, selectors, str(customer_id))
        # En algunos builds, Enter ayuda a disparar el filtro.
        try:
            self._press_any(page, selectors, "Enter")
        except Exception:
            pass

    def _fast_search_pick_client(self, page: Page, customer_id: str) -> bool:
        wanted_digits = self._id_key(str(customer_id))
        if not wanted_digits:
            return False

        # Esperar a que el contenedor de resultados esté visible.
        try:
            page.locator("css=#fast_search_result").first.wait_for(state="visible", timeout=10_000)
        except Exception:
            return False

        # Preferir explícitamente la opción de perfil cuyo primer renglón empieza por "Cliente".
        # Esto evita seleccionar "Servicio de Internet" (que también incluye texto "Cliente").
        rows = page.locator("css=#fast_search_result > tr")

        deadline = time.monotonic() + 12.0
        while True:
            if time.monotonic() > deadline:
                return False
            try:
                if rows.count() > 0:
                    break
            except Exception:
                pass
            time.sleep(0.2)

        def _first_line(txt: str) -> str:
            t = "\n".join([ln.strip() for ln in (txt or "").splitlines() if ln.strip()])
            if not t:
                return ""
            return t.split("\n", 1)[0].strip()

        best_idx: int | None = None
        best_score = -1
        row_count = 0
        try:
            row_count = rows.count()
        except Exception:
            row_count = 0

        for i in range(min(row_count, 30)):
            try:
                txt = rows.nth(i).inner_text() or ""
            except Exception:
                continue

            compact = " ".join(txt.split())
            if wanted_digits not in compact:
                continue

            first = self._norm_text(_first_line(txt))
            all_norm = self._norm_text(compact)

            score = 0
            # match por ID
            score += 10
            # preferir que la primera línea sea "Cliente ..."
            if first.startswith("cliente"):
                score += 100
            # penalizar entradas que son de servicios/documentos/etc.
            if "servicio" in first or "invoice" in first or "recibo" in first or "pago" in first or "documento" in first:
                score -= 50
            # pequeño plus si explícitamente contiene "cliente:" en la primera línea
            if "cliente:" in first:
                score += 20
            # si el texto completo contiene "servicio de internet" también penaliza
            if "servicio de internet" in all_norm:
                score -= 25

            if score > best_score:
                best_score = score
                best_idx = i

        if best_idx is None:
            # Fallback: intenta click por selectores (menos confiable, pero mejor que fallar).
            candidates = [
                f"css=#fast_search_result > tr:has-text('Cliente:'):has-text('{wanted_digits}') a",
                f"css=#fast_search_result > tr:has-text('Cliente:'):has-text('{wanted_digits}') td",
                f"css=#fast_search_result > tr:has-text('{wanted_digits}') a",
                f"css=#fast_search_result > tr:has-text('{wanted_digits}') td",
            ]
            try:
                self._click_any(page, candidates)
                return True
            except Exception:
                return False

        try:
            rows.nth(best_idx).locator("td").first.click()
            return True
        except Exception:
            try:
                rows.nth(best_idx).click()
                return True
            except Exception:
                return False

    def _fast_search_pick_ticket(self, page: Page, ticket_id: str) -> bool:
        wanted_digits = self._id_key(str(ticket_id))
        if not wanted_digits:
            return False

        try:
            page.locator("css=#fast_search_result").first.wait_for(state="visible", timeout=10_000)
        except Exception:
            return False

        rows = page.locator("css=#fast_search_result > tr")

        # Esperar a que haya al menos UNA fila que realmente corresponda al ticket.
        deadline = time.monotonic() + 15.0
        wanted_re = re.compile(rf"(?<!\d){re.escape(wanted_digits)}(?!\d)")

        def _row_has_exact_ticket_id(row_idx: int) -> bool:
            row = rows.nth(row_idx)

            # Preferir href con ?id=<wanted> exacto (evita falsos positivos tipo 313118 dentro de 2313118)
            try:
                links = row.locator("css=a")
                link_count = min(links.count(), 8)
            except Exception:
                link_count = 0

            for j in range(link_count):
                try:
                    href = (links.nth(j).get_attribute("href") or "").strip()
                except Exception:
                    href = ""
                if not href:
                    continue
                if "ticket" not in href and "tickets" not in href:
                    continue
                m = re.search(r"[?&]id=(\d+)", href)
                if m and m.group(1) == wanted_digits:
                    return True

            try:
                txt = row.inner_text() or ""
            except Exception:
                txt = ""
            return bool(wanted_re.search(" ".join(txt.split())))

        while True:
            if time.monotonic() > deadline:
                return False

            try:
                row_count = rows.count()
            except Exception:
                row_count = 0

            found = False
            for i in range(min(row_count, 60)):
                try:
                    if _row_has_exact_ticket_id(i):
                        found = True
                        break
                except Exception:
                    continue

            if found:
                break
            time.sleep(0.25)

        # Prioridad absoluta: si existe un link de ticket con href ?id=<wanted>, clickealo directamente.
        # Esto evita que se seleccione la primera opción cuando no corresponde.
        try:
            row_count = rows.count()
        except Exception:
            row_count = 0

        for i in range(min(row_count, 60)):
            row = rows.nth(i)
            try:
                links = row.locator("css=a")
                link_count = min(links.count(), 10)
            except Exception:
                link_count = 0

            for j in range(link_count):
                try:
                    href = (links.nth(j).get_attribute("href") or "").strip()
                except Exception:
                    href = ""
                if not href:
                    continue
                if "ticket" not in href and "tickets" not in href:
                    continue
                m = re.search(r"[?&]id=(\d+)", href)
                if m and m.group(1) == wanted_digits:
                    try:
                        links.nth(j).click()
                    except Exception:
                        try:
                            links.nth(j).click(force=True)
                        except Exception:
                            links.nth(j).evaluate("el => el.click()")
                    return True

        def _first_line(txt: str) -> str:
            t = "\n".join([ln.strip() for ln in (txt or "").splitlines() if ln.strip()])
            if not t:
                return ""
            return t.split("\n", 1)[0].strip()

        best_idx: int | None = None
        best_score = -1
        try:
            row_count = rows.count()
        except Exception:
            row_count = 0

        for i in range(min(row_count, 40)):
            try:
                txt = rows.nth(i).inner_text() or ""
            except Exception:
                continue

            compact = " ".join(txt.split())
            # Evitar coincidencias parciales dentro de otros números largos.
            if not wanted_re.search(compact):
                continue

            first = self._norm_text(_first_line(txt))
            all_norm = self._norm_text(compact)

            score = 0
            score += 10  # match por ID
            if "ticket" in all_norm:
                score += 80
            if first.startswith("ticket") or first.startswith("closed ticket") or first.startswith("open ticket"):
                score += 50

            # penalizar entradas que no son tickets
            if "cliente" in first or "cliente" in all_norm:
                score -= 30
            if "pago" in first or "invoice" in first or "recibo" in first or "documento" in first:
                score -= 40

            if score > best_score:
                best_score = score
                best_idx = i

        if best_idx is None:
            # Fallback explícito del usuario (puede variar, pero sirve para pruebas)
            candidates = [
                f"css=#fast_search_result > tr:has-text('{wanted_digits}'):has-text('ticket') td",
                f"css=#fast_search_result > tr:has-text('{wanted_digits}'):has-text('Ticket') td",
                "css=#fast_search_result > tr:nth-child(2) > td",
                "xpath=//*[@id='fast_search_result']/tr[2]/td",
                "xpath=/html/body/div[2]/div[5]/div/div[2]/div/div[2]/div/table/tr[2]/td",
            ]
            try:
                self._click_any(page, candidates)
                return True
            except Exception:
                return False

        # Si hay un <a href="...id=<wanted>"> dentro de la fila ganadora, clickealo.
        row = rows.nth(best_idx)
        try:
            links = row.locator("css=a")
            link_count = min(links.count(), 10)
        except Exception:
            link_count = 0

        for j in range(link_count):
            try:
                href = (links.nth(j).get_attribute("href") or "").strip()
            except Exception:
                href = ""
            if not href:
                continue
            m = re.search(r"[?&]id=(\d+)", href)
            if m and m.group(1) == wanted_digits:
                try:
                    links.nth(j).click()
                    return True
                except Exception:
                    try:
                        links.nth(j).click(force=True)
                        return True
                    except Exception:
                        pass

        try:
            rows.nth(best_idx).locator("td").first.click()
            return True
        except Exception:
            try:
                rows.nth(best_idx).click()
                return True
            except Exception:
                return False

    def _ensure_customer_info_tab(self, page: Page) -> None:
        """Fuerza la pestaña 'Información' del perfil para que los campos existan/estén visibles."""
        selectors = [
            "css=a:has-text('Información')",
            "css=li:has-text('Información') a",
            "text=Información",
            "css=a:has-text('Informacion')",
            "text=Informacion",
        ]

        # Best-effort: si ya está activa, click no hace daño.
        try:
            self._click_any(page, selectors)
        except Exception:
            pass

    def _wait_visible_any(self, page: Page, selectors: list[str], timeout_ms: int) -> bool:
        """Espera a que alguno de los selectores sea visible en page o en el scope (iframe)."""
        start = time.monotonic()
        while True:
            if (time.monotonic() - start) * 1000.0 > timeout_ms:
                return False

            # 1) scope (iframe) si aplica
            try:
                scope = self._get_scope(page)
                for sel in selectors:
                    try:
                        loc = scope.locator(sel).first
                        if loc.count() > 0 and loc.is_visible():
                            return True
                    except Exception:
                        continue
            except Exception:
                pass

            # 2) page raíz
            for sel in selectors:
                try:
                    loc = page.locator(sel).first
                    if loc.count() > 0 and loc.is_visible():
                        return True
                except Exception:
                    continue

            time.sleep(0.25)

    def _extract_socio_from_profile(self, page: Page) -> str:
        """Extrae el valor mostrado en el Select2 de 'Socio' dentro del perfil del cliente."""

        def _try(scope: LocatorScope) -> str:
            selectors = [
                "css=#admin_customers_view_form span[id^='select2-customers-partner_id-id-'][id$='-container']",
                "css=span[id^='select2-customers-partner_id-id-'][id$='-container']",
                "css=[id^='select2-customers-partner_id-id-'][id$='-container']",
            ]
            for sel in selectors:
                try:
                    loc = scope.locator(sel).first
                    if loc.count() == 0:
                        continue
                    loc.wait_for(state="visible", timeout=5_000)
                    txt = (loc.inner_text() or "").strip()
                    txt = " ".join(txt.split())
                    if txt:
                        return txt
                except Exception:
                    continue
            return ""

        try:
            v = _try(self._get_scope(page))
            if v:
                return v
        except Exception:
            pass
        return _try(page)

    def _extract_socio_from_profile_id(self, page: Page, customer_id: str) -> str:
        cid = self._id_key(customer_id)
        if not cid:
            return ""
        # Selector exacto para el ID actual
        exact = f"css=#select2-customers-partner_id-id-{cid}-container"
        try:
            loc = page.locator(exact).first
            if loc.count() > 0:
                loc.wait_for(state="visible", timeout=5_000)
                txt = (loc.inner_text() or "").strip()
                return " ".join(txt.split())
        except Exception:
            pass
        return self._extract_socio_from_profile(page)

    def _extract_res_urb_from_profile(self, page: Page) -> str:
        """Extrae el valor del input 'Residencia/Urbanización' dentro del perfil del cliente."""

        def _try(scope: LocatorScope) -> str:
            selectors = [
                "css=#admin_customers_view_form input[id^='customers-additional_attributes-res_urb-id-']",
                "css=input[id^='customers-additional_attributes-res_urb-id-']",
                "css=[id^='customers-additional_attributes-res_urb-id-']",
            ]
            for sel in selectors:
                try:
                    loc = scope.locator(sel).first
                    if loc.count() == 0:
                        continue
                    loc.wait_for(state="visible", timeout=5_000)
                    # input_value es lo más confiable para inputs
                    try:
                        v = loc.input_value(timeout=2_000)
                    except Exception:
                        v = (loc.get_attribute("value") or "").strip()
                    v = " ".join((v or "").split())
                    if v:
                        return v
                except Exception:
                    continue
            return ""

        try:
            v = _try(self._get_scope(page))
            if v:
                return v
        except Exception:
            pass
        return _try(page)

    def _extract_res_urb_from_profile_id(self, page: Page, customer_id: str) -> str:
        cid = self._id_key(customer_id)
        if not cid:
            return ""
        exact = f"css=#customers-additional_attributes-res_urb-id-{cid}"
        try:
            loc = page.locator(exact).first
            if loc.count() > 0:
                loc.wait_for(state="visible", timeout=5_000)
                try:
                    v = loc.input_value(timeout=2_000)
                except Exception:
                    v = (loc.get_attribute("value") or "").strip()
                return " ".join((v or "").split())
        except Exception:
            pass
        return self._extract_res_urb_from_profile(page)

    def _wait_customer_profile_loaded_for_id(self, page: Page, customer_id: str, timeout_ms: int) -> bool:
        """Espera que el perfil abierto corresponda al ID buscado.

        Evita el caso donde #admin_customers_view_form sigue visible del cliente anterior.
        """
        cid = self._id_key(customer_id)
        if not cid:
            return False

        # 1) URL (lo más fuerte cuando el routing es /customers/view?id=...)
        try:
            page.wait_for_function(
                f"() => String(window.location.href || '').includes('customers/view') && String(window.location.href || '').includes('id={cid}')",
                timeout=timeout_ms,
            )
            return True
        except Exception:
            pass

        # 2) Fallback: existencia de campos con sufijo del ID (selectores dinámicos)
        selectors = [
            f"css=#select2-customers-partner_id-id-{cid}-container",
            f"css=#customers-additional_attributes-res_urb-id-{cid}",
        ]
        return self._wait_visible_any(page, selectors, timeout_ms=timeout_ms)

    def _extract_value_by_label(self, page: Page, label_candidates: list[str]) -> str:
        # Intenta en scope (iframe) y luego en la página principal.
        def _try_in_scope(scope: LocatorScope) -> str:
            for root_sel in ("css=body", "css=html"):  # elemento ancla
                try:
                    root = scope.locator(root_sel).first
                    if root.count() == 0:
                        continue
                    val = root.evaluate(
                        r"""(el, labels) => {
                            const doc = el && el.ownerDocument ? el.ownerDocument : document;
                            const clean = (s) => String(s || '').replace(/\s+/g,' ').trim();
                            const norm = (s) => clean(s)
                                .toLowerCase()
                                .normalize('NFD')
                                .replace(/\p{Diacritic}/gu,'');
                            const wants = (labels || []).map(norm).filter(Boolean);
                            if (!wants.length) return '';

                            const nodes = Array.from(doc.querySelectorAll('th,td,dt,dd,label,div,span'));
                            for (const node of nodes) {
                                const t = norm(node.textContent || '');
                                if (!t) continue;
                                let hit = false;
                                for (const w of wants) {
                                    if (t === w || t === w + ':' || t.startsWith(w + ':') || t.includes(w)) {
                                        hit = true;
                                        break;
                                    }
                                }
                                if (!hit) continue;

                                // table row: <tr><th>Label</th><td>Value</td></tr>
                                const tr = node.closest('tr');
                                if (tr) {
                                    const tds = Array.from(tr.querySelectorAll('td'));
                                    if (tds.length) {
                                        const v = clean(tds[tds.length - 1].innerText || tds[tds.length - 1].textContent);
                                        if (v && norm(v) !== t) return v;
                                    }
                                }

                                // dl: <dt>Label</dt><dd>Value</dd>
                                if (node.tagName && node.tagName.toLowerCase() === 'dt') {
                                    const dd = node.nextElementSibling;
                                    if (dd && dd.tagName && dd.tagName.toLowerCase() === 'dd') {
                                        const v = clean(dd.innerText || dd.textContent);
                                        if (v) return v;
                                    }
                                }

                                // label for=input
                                const forId = node.getAttribute && node.getAttribute('for');
                                if (forId) {
                                    const inp = doc.getElementById(forId);
                                    if (inp) {
                                        const v = clean(inp.value || inp.textContent);
                                        if (v) return v;
                                    }
                                }

                                // next sibling
                                const sib = node.nextElementSibling;
                                if (sib) {
                                    const v = clean(sib.innerText || sib.textContent);
                                    if (v) return v;
                                }
                            }
                            return '';
                        }""",
                        label_candidates,
                    )
                    if isinstance(val, str) and val.strip():
                        return " ".join(val.split())
                except Exception:
                    continue
            return ""

        # 1) intentar en el iframe scope (si aplica)
        try:
            scope = self._get_scope(page)
            v = _try_in_scope(scope)
            if v:
                return v
        except Exception:
            pass

        # 2) fallback en la page raíz
        try:
            return _try_in_scope(page)
        except Exception:
            return ""

    def _upsert_customer_minimal(self, excel_path: str, *, customer_id: str, socio: str, residencia: str) -> None:
        wb = self._open_or_create_workbook(excel_path)
        ws = self._get_or_create_sheet(wb, "Datos Clientes")

        # asegurar headers mínimos
        if ws.max_row < 1:
            ws.append(["ID", "Socio", "Residencia/Urbanización"])
        if ws.max_row == 1 and (ws.cell(row=1, column=1).value is None):
            ws.append(["ID", "Socio", "Residencia/Urbanización"])

        headers = [str(ws.cell(row=1, column=c).value or "").strip() for c in range(1, ws.max_column + 1)]
        norm_to_col: dict[str, int] = {}
        for idx0, h in enumerate(headers):
            key = self._norm_text(h)
            if key and key not in norm_to_col:
                norm_to_col[key] = idx0 + 1

        def _ensure_col(name: str) -> int:
            key = self._norm_text(name)
            col = norm_to_col.get(key)
            if col:
                return col
            # crear al final
            col = ws.max_column + 1
            ws.cell(row=1, column=col).value = name
            norm_to_col[key] = col
            return col

        c_id = _ensure_col("ID")
        c_socio = _ensure_col("Socio")
        c_res = _ensure_col("Residencia/Urbanización")

        cid_key = self._id_key(customer_id)
        if not cid_key:
            return

        # buscar fila existente
        target_row = None
        for r in range(2, ws.max_row + 1):
            existing = self._id_key(ws.cell(row=r, column=c_id).value)
            if existing == cid_key:
                target_row = r
                break

        if target_row is None:
            target_row = ws.max_row + 1
            ws.cell(row=target_row, column=c_id).value = cid_key

        if socio:
            ws.cell(row=target_row, column=c_socio).value = socio
        if residencia:
            ws.cell(row=target_row, column=c_res).value = residencia

        wb.save(excel_path)

    def _do_enrich_missing(self, page: Page) -> None:
        excel_path = self._enrich_excel_path
        if not excel_path:
            self._message("Búsqueda/enriquecimiento: no se recibió ruta del Excel.")
            return
        if self._enrich_running:
            self._message("Búsqueda/enriquecimiento: ya está en progreso.")
            return

        self._enrich_running = True
        try:
            if not os.path.exists(excel_path):
                self._message(f"Búsqueda/enriquecimiento: no existe el archivo: {excel_path}")
                return

            ids = self._read_missing_ids(excel_path)
            if not ids:
                self._message("Búsqueda/enriquecimiento: no hay IDs para buscar.")
                return

            self._message(f"Búsqueda/enriquecimiento: {len(ids)} IDs únicos por buscar...")

            found = 0
            not_found = 0
            for idx, cid in enumerate(ids, start=1):
                if self._shutdown_event.is_set():
                    break

                try:
                    self._message(f"[{idx}/{len(ids)}] Buscando cliente ID {cid}...")
                    before_url = ""
                    try:
                        before_url = page.url
                    except Exception:
                        before_url = ""
                    self._fast_search_fill(page, cid)
                    page.wait_for_timeout(300)
                    ok = self._fast_search_pick_client(page, cid)
                    if not ok:
                        self._message(f"[{idx}/{len(ids)}] ID {cid}: no se pudo seleccionar el perfil (no encontrado o ambiguo).")
                        not_found += 1
                        continue

                    # Esperar navegación real (URL cambió o coincide con ID)
                    try:
                        if before_url:
                            page.wait_for_function(
                                "(u) => String(window.location.href || '') !== String(u || '')",
                                arg=before_url,
                                timeout=8_000,
                            )
                    except Exception:
                        pass

                    # Esperar que cargue el perfil CORRECTO para este ID.
                    ready = self._wait_customer_profile_loaded_for_id(page, cid, timeout_ms=25_000)
                    if not ready:
                        self._message(f"[{idx}/{len(ids)}] ID {cid}: timeout esperando el perfil. Saltando...")
                        not_found += 1
                        continue

                    # Asegurar la pestaña Información (algunas navegaciones abren #services u otra sección)
                    self._ensure_customer_info_tab(page)
                    page.wait_for_timeout(300)

                    # Esperar que los campos específicos del ID estén visibles (si existen)
                    self._wait_visible_any(
                        page,
                        [
                            f"css=#select2-customers-partner_id-id-{self._id_key(cid)}-container",
                            f"css=#customers-additional_attributes-res_urb-id-{self._id_key(cid)}",
                        ],
                        timeout_ms=8_000,
                    )

                    # Extraer los 2 campos cruciales (selectores que dependen del ID)
                    socio = self._extract_socio_from_profile_id(page, cid)
                    residencia = self._extract_res_urb_from_profile_id(page, cid)

                    # Fallback por texto/labels si algo cambió en el DOM
                    if not socio:
                        socio = self._extract_value_by_label(page, ["Socio", "Partner", "Socio:"])
                    if not residencia:
                        residencia = self._extract_value_by_label(
                            page,
                            [
                                "Residencia/Urbanización",
                                "Residencia",
                                "Urbanización",
                                "Urbanizacion",
                                "Dirección",
                                "Direccion",
                                "Address",
                            ],
                        )

                    if socio or residencia:
                        self._upsert_customer_minimal(
                            excel_path,
                            customer_id=cid,
                            socio=socio,
                            residencia=residencia,
                        )
                        found += 1
                        self._message(
                            f"[{idx}/{len(ids)}] ID {cid}: OK (Socio='{socio or '-'}', Residencia='{residencia or '-'}')."
                        )
                        page.wait_for_timeout(350)
                    else:
                        # No pudimos leer campos, pero el cliente abrió
                        self._message(f"[{idx}/{len(ids)}] ID {cid}: perfil abierto pero no se pudieron leer los campos.")
                        not_found += 1
                        page.wait_for_timeout(250)
                except PermissionError:
                    raise
                except Exception:
                    self._message(f"[{idx}/{len(ids)}] ID {cid}: error inesperado durante búsqueda/extracción.")
                    not_found += 1
                    page.wait_for_timeout(250)

                if idx % 10 == 0 or idx == len(ids):
                    self._message(
                        f"Búsqueda/enriquecimiento: progreso {idx}/{len(ids)}. Con datos: {found}, sin datos: {not_found}."
                    )

            self._message("Búsqueda/enriquecimiento: relanzando merge para actualizar 'Datos Completos'...")
            total, joined, nf = merge_tickets_customers(excel_path)
            self._message(f"OK: merge actualizado. Tickets: {total}, coincidencias: {joined}, no encontrados: {nf}.")
        except PermissionError:
            self._message(
                "Búsqueda/enriquecimiento: no pude abrir/guardar el Excel. "
                "Cierra 'output/Datos Splynx.xlsx' en Excel y vuelve a intentar."
            )
        except Exception as exc:
            self._message(f"Búsqueda/enriquecimiento: error: {exc}")
        finally:
            self._enrich_running = False

    def _do_extract(self, page: Page, table_key: str, mode: str = "auto") -> None:
        table_cfg = self._config.tables.get(table_key)
        if not table_cfg:
            self._message(f"No hay config para {table_key} en config.json")
            return

        scope = self._get_scope(page)

        try:
            mode = str(mode or "auto").lower()

            # Pasos previos configurables (para navegación en la UI antes de extraer)
            steps: list[Any] = list(table_cfg.get("steps", []))

            if table_key == "table1" and mode == "manual":
                # En modo manual el bot solo navega hasta Tickets -> List (y ajusta Acceso rápido).
                # Luego el usuario coloca filtros y presiona Aplicar.
                nav_steps: list[Any] = [
                    # Tickets
                    "css=body > div.splynx-wrapper > div.main > nav > div > div > div.sidebar-menu > div > div.menu-list > div:nth-child(4) > div > a||xpath=/html/body/div[2]/div[3]/nav/div/div/div[2]/div/div[1]/div[4]/div/a",
                    # List
                    "css=body > div.splynx-wrapper > div.main > nav > div > div > div.sidebar-menu > div > div.menu-list > div:nth-child(4) > div > div > div:nth-child(2) > div > a||xpath=/html/body/div[2]/div[3]/nav/div/div/div[2]/div/div[1]/div[4]/div/div/div[2]/div/a",
                    # Acceso rápido
                    "css=#select2-admin_support_tickets_opened_filter_quick_access-container||xpath=//*[@id='select2-admin_support_tickets_opened_filter_quick_access-container']",
                    "css=li[id^='select2-admin_support_tickets_opened_filter_quick_access-result-']:has-text('All tickets')||xpath=//li[starts-with(@id,'select2-admin_support_tickets_opened_filter_quick_access-result-')][contains(.,'All tickets')]||text=All tickets||xpath=/html/body/span/span/span[2]/ul/li[4]",
                ]

                self._message("Tabla 1 (manual): navegando a Tickets > List...")
                for step in nav_steps:
                    self._run_step(page, scope, step)
                    page.wait_for_timeout(600)

                # La navegación puede cambiar el iframe; recalcular scope.
                scope = self._get_scope(page)

                self._message(
                    "Tabla 1 (manual): abre 'Filter', coloca tus filtros y presiona 'Aplicar'. "
                    "La extracción empezará DESPUÉS de ese clic."
                )

                start_marker = self._tickets_first_row_marker(scope)
                self._wait_for_apply_click(scope, timeout_s=900.0)
                # Esperar la recarga tras Aplicar (si no cambia, igual continuamos tras un pequeño delay)
                self._wait_for_tickets_reload_after_apply(page, scope, start_marker=start_marker, timeout_s=180.0)
                output_xlsx = os.path.join("output", "Datos Splynx.xlsx")
                os.makedirs(os.path.dirname(output_xlsx) or ".", exist_ok=True)
                self._message("Extrayendo datos de Tickets a Excel...")
                page.wait_for_timeout(800)
                extract_tickets_to_excel(scope, output_xlsx=output_xlsx, sheet_name="Datos de Tickets")
                self._message(f"OK: exportado a {output_xlsx}")
                return

            if table_key == "table1" and not steps:
                # Defaults (Tickets -> List) + filtros nuevos
                steps = [
                    # Tickets
                    "css=body > div.splynx-wrapper > div.main > nav > div > div > div.sidebar-menu > div > div.menu-list > div:nth-child(4) > div > a||xpath=/html/body/div[2]/div[3]/nav/div/div/div[2]/div/div[1]/div[4]/div/a",
                    # List
                    "css=body > div.splynx-wrapper > div.main > nav > div > div > div.sidebar-menu > div > div.menu-list > div:nth-child(4) > div > div > div:nth-child(2) > div > a||xpath=/html/body/div[2]/div[3]/nav/div/div/div[2]/div/div[1]/div[4]/div/div/div[2]/div/a",
                    # Quick access dropdown (Acceso Rápido)
                    "css=#select2-admin_support_tickets_opened_filter_quick_access-container||xpath=//*[@id='select2-admin_support_tickets_opened_filter_quick_access-container']",
                    # All tickets option (id suele cambiar; preferimos texto/xpath)
                    "css=li[id^='select2-admin_support_tickets_opened_filter_quick_access-result-']:has-text('All tickets')||xpath=//li[starts-with(@id,'select2-admin_support_tickets_opened_filter_quick_access-result-')][contains(.,'All tickets')]||text=All tickets||xpath=/html/body/span/span/span[2]/ul/li[4]",
                    # Filter button
                    "css=#content > div > div.splynx-top-nav > div.filters-nav > div > div:nth-child(6) > button||xpath=//*[@id='content']/div/div[1]/div[2]/div/div[6]/button||xpath=/html/body/div[2]/div[3]/div[1]/div/div/div[1]/div[2]/div/div[6]/button||text=Filter",

                    # Condition
                    "css=#select2-admin_support_tickets_opened_search_widget_condition-container||xpath=//*[@id='select2-admin_support_tickets_opened_search_widget_condition-container']",
                    # All option (id dinámico)
                    "css=li[id^='select2-admin_support_tickets_opened_search_widget_condition-result-']:has-text('All')||css=li[id^='select2-admin_support_tickets_opened_search_widget_condition-result-']:has-text('Todos')||xpath=//li[starts-with(@id,'select2-admin_support_tickets_opened_search_widget_condition-result-')][contains(.,'All') or contains(.,'Todos')]||xpath=/html/body/span/span/span[2]/ul/li[3]||text=All||text=Todos",

                    # Group
                    "css=#select2-admin_support_tickets_opened_search_widget_group_id-container||xpath=//*[@id='select2-admin_support_tickets_opened_search_widget_group_id-container']",
                    # Buscar "Cualquiera"
                    {
                        "action": "fill",
                        "selector": "css=body > span > span > span.select2-search.select2-search--dropdown > input||xpath=/html/body/span/span/span[1]/input||xpath=//*[@id='opened-page']/body/span/span/span[1]/input||xpath=//*[@id='opened--view-page']/body/span/span/span[1]/input",
                        "text": "Cualquiera",
                    },
                    # Seleccionar opción "Cualquiera" (o resaltada)
                    "css=#select2-admin_support_tickets_opened_search_widget_group_id-results li.select2-results__option--highlighted||css=#select2-admin_support_tickets_opened_search_widget_group_id-results li.select2-results__option:has-text('Cualquiera')||css=#select2-admin_support_tickets_opened_search_widget_group_id-results li.select2-results__option:has-text('Any')||xpath=//*[@id='select2-admin_support_tickets_opened_search_widget_group_id-results']/li[contains(.,'Cualquiera') or contains(.,'Any')]||xpath=/html/body/span/span/span[2]/ul/li[16]",

                    # Socio
                    "css=#select2-admin_support_tickets_opened_search_widget_partner_id-container||xpath=//*[@id='select2-admin_support_tickets_opened_search_widget_partner_id-container']",
                    # Buscar "Cualquiera"
                    {
                        "action": "fill",
                        "selector": "css=body > span > span > span.select2-search.select2-search--dropdown > input||xpath=/html/body/span/span/span[1]/input||xpath=//*[@id='opened-page']/body/span/span/span[1]/input||xpath=//*[@id='opened--view-page']/body/span/span/span[1]/input",
                        "text": "Cualquiera",
                    },
                    # Seleccionar opción "Cualquiera" (o resaltada)
                    "css=#select2-admin_support_tickets_opened_search_widget_partner_id-results li.select2-results__option--highlighted||css=#select2-admin_support_tickets_opened_search_widget_partner_id-results li.select2-results__option:has-text('Cualquiera')||css=#select2-admin_support_tickets_opened_search_widget_partner_id-results li.select2-results__option:has-text('Any')||xpath=//ul[@id='select2-admin_support_tickets_opened_search_widget_partner_id-results']/li[contains(.,'Cualquiera') or contains(.,'Any')]||xpath=/html/body/span/span/span[2]/ul/li[1]",

                    # Period: llenar rango de fechas (desde inicio de mes hasta hoy)
                    {
                        "action": "wait_nonempty",
                        "selector": "css=#admin_support_tickets_opened_search_widget_created_at||xpath=//*[@id='admin_support_tickets_opened_search_widget_created_at']||xpath=/html/body/div[2]/div[3]/div[1]/div/div/div[2]/div/div[2]/div/div/form/div/div[2]/div/div/input[1]",
                        "timeout_ms": 300000
                    },
                    # Forzar commit/blur del input de Period
                    {
                        "action": "press",
                        "selector": "css=#admin_support_tickets_opened_search_widget_created_at||xpath=//*[@id='admin_support_tickets_opened_search_widget_created_at']||xpath=/html/body/div[2]/div[3]/div[1]/div/div/div[2]/div/div[2]/div/div/form/div/div[2]/div/div/input[1]",
                        "key": "Tab",
                    },
                    # Esperar a que el botón Aplicar esté habilitado
                    {
                        "action": "wait_enabled",
                        "selector": "css=#admin_support_tickets_opened_search_block > div > div > div > button.btn.btn-primary.ms-4.advanced-filter-apply-button||css=#admin_support_tickets_opened_search_block button.btn.btn-primary.advanced-filter-apply-button||css=button.advanced-filter-apply-button:has-text('Aplicar')||css=button.advanced-filter-apply-button:has-text('Apply')||xpath=//*[@id='admin_support_tickets_opened_search_block']/div/div/div/button[2]||xpath=/html/body/div[2]/div[3]/div[1]/div/div/div[2]/div/div[2]/div/div/div/button[2]||text=Aplicar||text=Apply",
                        "timeout_ms": 120000,
                    },
                    # Aplicar filtros
                    "css=button.advanced-filter-apply-button:has-text('Aplicar')||css=button.advanced-filter-apply-button:has-text('Apply')||css=#admin_support_tickets_opened_search_block button.advanced-filter-apply-button||css=#admin_support_tickets_opened_search_block > div > div > div > button.btn.btn-primary.ms-4.advanced-filter-apply-button||xpath=//*[@id='admin_support_tickets_opened_search_block']/div/div/div/button[2]||xpath=/html/body/div[2]/div[3]/div[1]/div/div/div[2]/div/div[2]/div/div/div/button[2]||text=Aplicar||text=Apply",
                ]

            if table_key == "table2" and not steps:
                # Defaults: Clientes -> Lista
                steps = [
                    "css=body > div.splynx-wrapper > div.main > nav > div > div > div.sidebar-menu > div > div.menu-list > div:nth-child(2) > div > a",
                    "css=body > div.splynx-wrapper > div.main > nav > div > div > div.sidebar-menu > div > div.menu-list > div:nth-child(2) > div > div > div:nth-child(2) > div > a||xpath=//*[@id='list-page']/body/div[2]/div[3]/nav/div/div/div[2]/div/div[1]/div[2]/div/div/div[2]/div/a||xpath=/html/body/div[2]/div[3]/nav/div/div/div[2]/div/div[1]/div[2]/div/div/div[2]/div/a||xpath=//*[@id='list-page']/body/div[2]/div[3]/nav/div/div/div[2]/div/div[1]/div[2]/div/a||xpath=/html/body/div[2]/div[3]/nav/div/div/div[2]/div/div[1]/div[2]/div/a",
                ]

            # Si hay steps, ejecutarlos en orden. Permitimos fallbacks por paso con separador "||".
            if steps:
                self._message(f"Ejecutando navegación previa para {table_key}...")
                for step in steps:
                    self._run_step(page, scope, step)
                    page.wait_for_timeout(600)

            if table_key == "table1":
                # Luego de completar filtros, exporta la tabla visible a Excel.
                output_xlsx = os.path.join("output", "Datos Splynx.xlsx")
                os.makedirs(os.path.dirname(output_xlsx) or ".", exist_ok=True)

                self._message("Extrayendo datos de Tickets a Excel...")
                # Pequeña espera para que el listado recargue tras Aplicar
                page.wait_for_timeout(1200)
                extract_tickets_to_excel(scope, output_xlsx=output_xlsx, sheet_name="Datos de Tickets")
                self._message(f"OK: exportado a {output_xlsx}")
                return

            if table_key == "table2":
                output_xlsx = os.path.join("output", "Datos Splynx.xlsx")
                os.makedirs(os.path.dirname(output_xlsx) or ".", exist_ok=True)

                self._message("Extrayendo datos de Clientes a Excel (solo esta página)...")
                page.wait_for_timeout(1200)
                extract_customers_to_excel(scope, output_xlsx=output_xlsx, sheet_name="Datos Clientes")
                self._message(f"OK: Clientes exportados a {output_xlsx} (hoja: Datos Clientes)")
                return

            selector = table_cfg.get("selector")
            output_csv = table_cfg.get("output_csv")

            if not selector or not output_csv:
                self._message(f"Config incompleta para {table_key}: falta selector u output_csv")
                return

            os.makedirs(os.path.dirname(output_csv) or ".", exist_ok=True)

            self._message(f"Extrayendo {table_key} usando selector: {selector}")
            extract_table_to_csv(scope, selector=selector, output_csv=output_csv)
            self._message(f"OK: {table_key} exportada a {output_csv}")
        except Exception as exc:
            self._message(f"Error extrayendo {table_key}: {exc}")

    def _tickets_first_row_marker(self, scope: LocatorScope) -> str:
        table_sel = "css=#admin_support_tickets_opened_list"
        try:
            first_row = scope.locator(f"{table_sel} tbody tr").first
            if first_row.count() == 0:
                return ""
            return " ".join(first_row.inner_text().strip().split())
        except Exception:
            return ""

    def _wait_for_apply_click(self, scope: LocatorScope, timeout_s: float) -> None:
        """Espera a que el usuario haga clic en el botón Aplicar (en el contexto correcto/iframe)."""
        selectors = [
            "css=#admin_support_tickets_opened_search_block > div > div > div > button.btn.btn-primary.ms-4.advanced-filter-apply-button",
            "css=#admin_support_tickets_opened_search_block button.btn.btn-primary.advanced-filter-apply-button",
            "css=button.advanced-filter-apply-button:has-text('Aplicar')",
            "css=button.advanced-filter-apply-button:has-text('Apply')",
            "xpath=//*[@id='admin_support_tickets_opened_search_block']/div/div/div/button[2]",
            "xpath=/html/body/div[2]/div[3]/div[1]/div/div/div[2]/div/div[2]/div/div/div/button[2]",
            "text=Aplicar",
            "text=Apply",
        ]

        start = time.monotonic()
        last_exc: Exception | None = None

        # Esperar a que el botón exista y esté visible (normalmente aparece al abrir Filter).
        apply_loc = None
        for sel in selectors:
            try:
                loc = scope.locator(sel).first
                loc.wait_for(state="visible", timeout=5_000)
                apply_loc = loc
                break
            except Exception as exc:
                last_exc = exc
                continue

        while apply_loc is None:
            if time.monotonic() - start > timeout_s:
                raise TimeoutError("Timeout esperando que aparezca el botón 'Aplicar'.")
            # reintentar encontrarlo
            for sel in selectors:
                try:
                    loc = scope.locator(sel).first
                    loc.wait_for(state="visible", timeout=2_000)
                    apply_loc = loc
                    break
                except Exception as exc:
                    last_exc = exc
                    continue
            if apply_loc is None:
                time.sleep(0.25)

        # Instalar listener una vez (en el frame donde vive el botón)
        try:
            apply_loc.evaluate(
                """el => {
                    try { window.__splynxApplyClicks = 0; } catch(e) {}
                    if (!el.__splynxBound) {
                        el.__splynxBound = true;
                        el.addEventListener('click', () => { window.__splynxApplyClicks = (window.__splynxApplyClicks || 0) + 1; }, true);
                    }
                }"""
            )
        except Exception:
            # Si no se puede instalar, igual intentamos detectar cambios de tabla luego.
            pass

        # Esperar el click
        while True:
            if time.monotonic() - start > timeout_s:
                raise TimeoutError("Timeout esperando clic en 'Aplicar'.")
            try:
                clicks = apply_loc.evaluate("() => window.__splynxApplyClicks || 0")
                if int(clicks) >= 1:
                    return
            except Exception:
                # si el botón desaparece/re-renderiza, re-encontrarlo
                apply_loc = None
                for sel in selectors:
                    try:
                        loc = scope.locator(sel).first
                        loc.wait_for(state="visible", timeout=1_000)
                        apply_loc = loc
                        break
                    except Exception:
                        continue
            time.sleep(0.25)

    def _wait_for_tickets_reload_after_apply(
        self,
        page: Page,
        scope: LocatorScope,
        timeout_s: float = 300.0,
        start_marker: str | None = None,
    ) -> None:
        """Espera una recarga de la tabla de tickets tras presionar Aplicar.

        Nota: algunos casos no muestran overlay 'processing' o la primera fila puede quedar igual;
        en ese caso no bloqueamos indefinidamente.
        """
        table_sel = "css=#admin_support_tickets_opened_list"

        if start_marker is None:
            start_marker = self._tickets_first_row_marker(scope)

        start = time.monotonic()
        seen_processing = False
        while True:
            if time.monotonic() - start > timeout_s:
                # Si nunca vimos processing ni cambió marker, igual dejamos seguir para no bloquear.
                return

            # Detectar overlay de processing (DataTables)
            try:
                proc = scope.locator("css=div.dataTables_processing").first
                if proc.count() > 0 and proc.is_visible():
                    seen_processing = True
                    time.sleep(0.25)
                    continue
            except Exception:
                pass

            # Si ya vimos processing, esperar a que haya filas (o estado vacío) y salir
            try:
                table = scope.locator(table_sel).first
                if table.count() > 0:
                    rows = table.locator("tbody tr")
                    if rows.count() > 0:
                        # estado vacío
                        if table.locator("tbody tr td.dataTables_empty").count() > 0:
                            return
                        # si cambió el primer row o al menos ya hay datos
                        if seen_processing:
                            return
                        if start_marker and self._tickets_first_row_marker(scope) != start_marker:
                            return
            except Exception:
                pass

            # Si el primer row cambia, asumimos recarga
            if start_marker and self._tickets_first_row_marker(scope) != start_marker:
                return

            time.sleep(0.25)
