from __future__ import annotations

import re
import unicodedata
from decimal import Decimal, InvalidOperation
from dataclasses import dataclass
from pathlib import Path
from typing import Dict, List, Tuple

from openpyxl.reader.excel import load_workbook


def _norm_header(s: str) -> str:
    s = (s or "").strip().lower()
    s = "".join(c for c in unicodedata.normalize("NFKD", s) if not unicodedata.combining(c))
    s = re.sub(r"\s+", " ", s)
    return s


def _digits(s: str) -> str:
    matches = re.findall(r"\d+", (s or ""))
    if not matches:
        return ""
    # Usa el grupo de dígitos más largo para evitar casos tipo "1.23E+05" donde el primer match sería "1".
    return max(matches, key=len)


_SCI_RE = re.compile(r"^\s*[-+]?\d+(?:\.\d+)?[eE][-+]?\d+\s*$")


def _id_key(value) -> str:
    """Normaliza IDs de Excel/strings a una llave comparable.

    - Acepta int/float/str.
    - Maneja notación científica (ej: '1.35921E+05').
    - Extrae dígitos (grupo más largo).
    - Quita ceros a la izquierda.
    """
    if value is None:
        return ""

    if isinstance(value, bool):
        return ""

    if isinstance(value, int):
        return str(value)

    if isinstance(value, float):
        try:
            return str(int(value))
        except Exception:
            pass

    s = str(value).strip()
    if not s:
        return ""

    if _SCI_RE.match(s):
        try:
            num = Decimal(s)
            s = str(int(num))
        except (InvalidOperation, ValueError):
            pass

    d = _digits(s)
    if not d:
        return ""
    d2 = d.lstrip("0")
    return d2 if d2 else "0"


@dataclass(frozen=True)
class SheetColumns:
    header_to_col: Dict[str, int]  # normalized header -> 1-based col index


def _get_sheet_columns(ws) -> SheetColumns:
    header_to_col: Dict[str, int] = {}
    for col_idx in range(1, ws.max_column + 1):
        val = ws.cell(row=1, column=col_idx).value
        if val is None:
            continue
        key = _norm_header(str(val))
        if key:
            header_to_col[key] = col_idx
    return SheetColumns(header_to_col=header_to_col)


def _require(cols: SheetColumns, name: str) -> int:
    key = _norm_header(name)
    if key not in cols.header_to_col:
        raise KeyError(f"No se encontró la columna '{name}' en el Excel.")
    return cols.header_to_col[key]


def _optional(cols: SheetColumns, name: str) -> int | None:
    key = _norm_header(name)
    return cols.header_to_col.get(key)


def merge_tickets_customers(
    excel_path: str | Path,
    tickets_sheet: str = "Datos de Tickets",
    customers_sheet: str = "Datos Clientes",
    output_sheet: str = "Datos Completos",
    not_found_sheet: str = "Datos no Encontrados",
    summary_sheet: str = "Resumen Merge",
) -> Tuple[int, int, int]:
    """Crea/rehace una hoja 'Datos Completos' haciendo join por ID.

    Join:
            - Tickets['Reporter ID'] (cuando Reporter type = 'customer')
            - Clientes['ID']

        Fallback (para tickets creados por admins u otros casos):
            - Si Reporter type != 'customer' o Reporter ID viene vacío, se intenta con Tickets['ID Cliente'].

    Escribe:
            - Columnas de Tickets (según lo solicitado):
                ID, Tema, Customer/Lead, Prioridad, Estado, Group, Tipo, Asignado a, Watching, Labels,
                Reporter, Reporter ID, Reporter type, ID Cliente, Incoming Customer, Hide, Task, Estrella,
                Creado (fecha y hora), Source, Actualizado (fecha y hora), Archive, Shareable, Note,
                Sub-tipo de Ticket, Categoria del Cierre, Promocion
    - Columnas de Clientes: Servicio usuario, Socio, Residencia/Urbanización

    Returns: (tickets_rows_total, rows_joined, rows_not_found)
    """

    excel_path = Path(excel_path)
    if not excel_path.exists():
        raise FileNotFoundError(f"No existe el archivo: {excel_path}")

    wb = load_workbook(excel_path)
    if tickets_sheet not in wb.sheetnames:
        raise KeyError(f"No existe la hoja '{tickets_sheet}'")
    if customers_sheet not in wb.sheetnames:
        raise KeyError(f"No existe la hoja '{customers_sheet}'")

    ws_t = wb[tickets_sheet]
    ws_c = wb[customers_sheet]

    t_cols = _get_sheet_columns(ws_t)
    c_cols = _get_sheet_columns(ws_c)

    # Tickets
    t_id = _require(t_cols, "ID")
    t_reporter_id = _optional(t_cols, "Reporter ID")
    t_reporter_type = _optional(t_cols, "Reporter type")
    t_id_cliente = _optional(t_cols, "ID Cliente")

    if not (isinstance(t_reporter_id, int) and t_reporter_id > 0) and not (
        isinstance(t_id_cliente, int) and t_id_cliente > 0
    ):
        raise KeyError(
            "No se encontró 'Reporter ID' ni 'ID Cliente' en 'Datos de Tickets' (necesito al menos una para comparar)."
        )

    ticket_out_cols: List[str] = [
        "ID",
        "Tema",
        "Customer/Lead",
        "Prioridad",
        "Estado",
        "Group",
        "Tipo",
        "Asignado a",
        "Watching",
        "Labels",
        "Reporter",
        "Reporter ID",
        "Reporter type",
        "ID Cliente",
        "Incoming Customer",
        "Hide",
        "Task",
        "Estrella",
        "Creado (fecha y hora)",
        "Source",
        "Actualizado (fecha y hora)",
        "Archive",
        "Shareable",
        "Note",
        "Sub-tipo de Ticket",
        "Categoria del Cierre",
        "Promocion",
    ]

    # Mapa de columnas de tickets para lectura (opcionales excepto ID/ID Cliente)
    t_col_idx: Dict[str, int | None] = {name: _optional(t_cols, name) for name in ticket_out_cols}
    t_col_idx["ID"] = t_id
    t_col_idx["Reporter ID"] = t_reporter_id
    t_col_idx["Reporter type"] = t_reporter_type
    t_col_idx["ID Cliente"] = t_id_cliente

    # Clientes
    c_id = _require(c_cols, "ID")
    c_servicio = _optional(c_cols, "Servicio usuario")
    c_socio = _require(c_cols, "Socio")
    c_res = _require(c_cols, "Residencia/Urbanización")

    # Build map: customer_id -> (servicio_usuario, socio, residencia)
    customer_map: Dict[str, Tuple[str, str, str]] = {}
    customer_row_by_id: Dict[str, int] = {}
    for r in range(2, ws_c.max_row + 1):
        cid = _id_key(ws_c.cell(row=r, column=c_id).value)
        if not cid:
            continue
        servicio = ""
        if isinstance(c_servicio, int) and c_servicio > 0:
            servicio = str(ws_c.cell(row=r, column=c_servicio).value or "").strip()
        socio = str(ws_c.cell(row=r, column=c_socio).value or "").strip()
        res = str(ws_c.cell(row=r, column=c_res).value or "").strip()
        customer_map[cid] = (servicio, socio, res)
        customer_row_by_id[cid] = r

    # Recreate output sheet
    if output_sheet in wb.sheetnames:
        wb.remove(wb[output_sheet])
    ws_o = wb.create_sheet(title=output_sheet)

    if not_found_sheet in wb.sheetnames:
        wb.remove(wb[not_found_sheet])
    ws_nf = wb.create_sheet(title=not_found_sheet)

    # Sheet de resumen/diagnóstico
    if summary_sheet in wb.sheetnames:
        wb.remove(wb[summary_sheet])
    ws_s = wb.create_sheet(title=summary_sheet)

    out_headers = [*ticket_out_cols, "Servicio usuario", "Socio", "Residencia/Urbanización"]
    ws_o.append(out_headers)

    nf_headers = [*ticket_out_cols]
    ws_nf.append(nf_headers)

    tickets_total = 0
    joined = 0
    not_found = 0
    not_found_counts: Dict[str, int] = {}

    # Stats básicos previos (ayuda a detectar exportación incompleta/filtrada)
    ws_s.append(["Métrica", "Valor"])
    ws_s.append(["Tickets (filas en hoja)", ws_t.max_row - 1])
    ws_s.append(["Clientes (filas en hoja)", ws_c.max_row - 1])
    ws_s.append(["Clientes (IDs únicos)", len(customer_map)])
    ws_s.append(["Join", "Tickets(Reporter ID/ID Cliente) -> Clientes(ID)"])

    def _safe_int(x: str) -> int | None:
        try:
            return int(x)
        except Exception:
            return None

    c_ints = [v for v in (_safe_int(k) for k in customer_map.keys()) if v is not None]
    if c_ints:
        ws_s.append(["Clientes ID min", min(c_ints)])
        ws_s.append(["Clientes ID max", max(c_ints)])

    # Diagnóstico
    used_reporter_id = 0
    used_id_cliente = 0
    reporter_type_customer = 0
    reporter_type_admin = 0
    reporter_type_other = 0
    blank_join_id = 0

    def _norm_type(v) -> str:
        s = str(v or "").strip().lower()
        return s

    for r in range(2, ws_t.max_row + 1):
        tickets_total += 1

        reporter_id_val = ""
        if isinstance(t_reporter_id, int) and t_reporter_id > 0:
            reporter_id_val = _id_key(ws_t.cell(row=r, column=t_reporter_id).value)

        reporter_type_val = ""
        if isinstance(t_reporter_type, int) and t_reporter_type > 0:
            reporter_type_val = _norm_type(ws_t.cell(row=r, column=t_reporter_type).value)

        if reporter_type_val == "customer":
            reporter_type_customer += 1
        elif reporter_type_val == "admin":
            reporter_type_admin += 1
        elif reporter_type_val:
            reporter_type_other += 1

        id_cliente_val = ""
        if isinstance(t_id_cliente, int) and t_id_cliente > 0:
            id_cliente_val = _id_key(ws_t.cell(row=r, column=t_id_cliente).value)

        # Selección de clave (prioriza Reporter ID si el reporter es customer)
        join_id = ""
        if reporter_type_val == "customer" and reporter_id_val:
            join_id = reporter_id_val
            used_reporter_id += 1
        elif id_cliente_val:
            join_id = id_cliente_val
            used_id_cliente += 1
        elif reporter_id_val:
            # Si no sabemos el tipo pero hay reporter_id, mejor intentar igual.
            join_id = reporter_id_val
            used_reporter_id += 1

        if not join_id:
            blank_join_id += 1
            not_found += 1
            not_found_counts[""] = not_found_counts.get("", 0) + 1
            nf_row: List = []
            for name in ticket_out_cols:
                col = t_col_idx.get(name)
                if isinstance(col, int) and col > 0:
                    nf_row.append(ws_t.cell(row=r, column=col).value)
                else:
                    nf_row.append("")
            ws_nf.append(nf_row)
            continue

        extra = customer_map.get(join_id)
        if not extra:
            not_found += 1
            not_found_counts[join_id] = not_found_counts.get(join_id, 0) + 1
            nf_row2: List = []
            for name in ticket_out_cols:
                col = t_col_idx.get(name)
                if isinstance(col, int) and col > 0:
                    nf_row2.append(ws_t.cell(row=r, column=col).value)
                else:
                    nf_row2.append("")
            ws_nf.append(nf_row2)
            continue

        servicio, socio, res = extra
        row: List = []
        for name in ticket_out_cols:
            col = t_col_idx.get(name)
            if isinstance(col, int) and col > 0:
                row.append(ws_t.cell(row=r, column=col).value)
            else:
                row.append("")

        row.extend([servicio, socio, res])
        ws_o.append(row)
        joined += 1

    # Completa resumen
    ws_s.append(["Tickets total (procesados)", tickets_total])
    ws_s.append(["Coincidencias (join)", joined])
    ws_s.append(["No encontrados", not_found])
    ws_s.append(["Tickets sin ID para comparar", blank_join_id])
    ws_s.append(["Join usando Reporter ID", used_reporter_id])
    ws_s.append(["Join usando ID Cliente (fallback)", used_id_cliente])
    if reporter_type_customer or reporter_type_admin or reporter_type_other:
        ws_s.append(["Reporter type = customer", reporter_type_customer])
        ws_s.append(["Reporter type = admin", reporter_type_admin])
        ws_s.append(["Reporter type = otros", reporter_type_other])

    # Lista de IDs no encontrados (top) con conteo
    ws_s.append([""])  # separador
    ws_s.append(["Top IDs no encontrados", "Conteo", "Ejemplo fila ticket", "Fila cliente (si existe)"])

    # Solo IDs con valor (evita la llave vacía)
    items = [(k, v) for k, v in not_found_counts.items() if k]
    items.sort(key=lambda kv: kv[1], reverse=True)
    # Limitar para no inflar el Excel
    for k, v in items[:200]:
        ws_s.append([k, v, "", customer_row_by_id.get(k, "")])

    wb.save(excel_path)
    return tickets_total, joined, not_found
