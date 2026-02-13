from __future__ import annotations

import csv
import re
import time
from dataclasses import dataclass
from pathlib import Path
from typing import List

from typing import Protocol

from openpyxl import Workbook
from openpyxl.reader.excel import load_workbook


class LocatorScope(Protocol):
    def locator(self, selector: str):  # playwright's Locator
        ...


def extract_table_to_csv(scope: LocatorScope, selector: str, output_csv: str) -> None:
    table = scope.locator(selector).first
    table.wait_for(state="visible")

    row_loc = table.locator("tr")
    row_count = row_loc.count()

    rows: List[List[str]] = []
    max_cols = 0

    for i in range(row_count):
        row = row_loc.nth(i)
        cell_loc = row.locator("th, td")
        cell_count = cell_loc.count()
        cells = []
        for j in range(cell_count):
            txt = cell_loc.nth(j).inner_text().strip()
            txt = " ".join(txt.split())
            cells.append(txt)
        max_cols = max(max_cols, len(cells))
        if cells:
            rows.append(cells)

    # Normaliza filas para que todas tengan el mismo número de columnas
    for r in rows:
        if len(r) < max_cols:
            r.extend([""] * (max_cols - len(r)))

    with open(output_csv, "w", newline="", encoding="utf-8-sig") as f:
        w = csv.writer(f)
        w.writerows(rows)


def extract_tickets_to_excel(
    scope: LocatorScope,
    output_xlsx: str,
    sheet_name: str = "Datos de Tickets",
    table_selector: str = "css=#admin_support_tickets_opened_list",
) -> None:
    table = scope.locator(table_selector).first
    table.wait_for(state="visible")

    def _norm_header(s: str) -> str:
        # Normaliza: minúsculas, espacios, y sin acentos (para matchear "actualización" vs "actualizacion").
        s = (s or "").strip().lower()
        try:
            import unicodedata

            s = "".join(
                c for c in unicodedata.normalize("NFKD", s) if not unicodedata.combining(c)
            )
        except Exception:
            pass
        s = " ".join(s.split())
        return s

    def _header_index_map() -> dict[str, int]:
        """Mapea header normalizado -> índice 0-based de la columna en la tabla."""
        table_loc = scope.locator(table_selector).first
        ths = table_loc.locator("thead tr th")
        try:
            texts = ths.all_inner_texts()
        except Exception:
            texts = []

        m: dict[str, int] = {}
        for idx, t in enumerate(texts):
            key = _norm_header(t)
            if key and key not in m:
                m[key] = idx
        return m

    header_map = _header_index_map()

    def _pick_idx(*aliases: str) -> int | None:
        for a in aliases:
            if not a:
                continue
            key = _norm_header(a)
            if key in header_map:
                return header_map[key]
        return None

    # Columnas a exportar (orden fijo). Se resuelven por texto del header.
    # Si alguna no existe/está oculta en la UI, se exporta vacía.
    col_map: list[tuple[str, int | None]] = [
        ("ID", _pick_idx("ID")),
        ("Tema", _pick_idx("Tema", "Subject")),
        ("Customer/Lead", _pick_idx("Customer / Lead", "Customer/Lead", "Customer", "Lead")),
        ("Prioridad", _pick_idx("Prioridad", "Priority")),
        ("Estado", _pick_idx("Estado", "Status")),
        ("Group", _pick_idx("Group", "Grupo")),
        ("Tipo", _pick_idx("Tipo", "Type")),
        ("Asignado a", _pick_idx("Asignado a", "Assignee", "Assigned to")),
        ("Watching", _pick_idx("Watching", "Watchers")),
        ("Labels", _pick_idx("Labels", "Etiquetas")),
        ("Reporter", _pick_idx("Reporter", "Reportero")),
        ("Reporter ID", _pick_idx("Reporter ID", "ReporterID")),
        ("Reporter type", _pick_idx("Reporter type", "Reporter Type")),
        (
            "ID Cliente",
            _pick_idx(
                "ID Cliente",
                "ID cliente",
                "ID de cliente",
                "ID de clientes",
                "ID del cliente",
                "Customer ID",
                "Client ID",
            ),
        ),
        ("Incoming Customer", _pick_idx("Incoming Customer")),
        ("Hide", _pick_idx("Hide")),
        ("Task", _pick_idx("Task", "Tarea")),
        ("Estrella", _pick_idx("Estrella", "Star")),
        (
            "Creado (fecha y hora)",
            _pick_idx(
                "Creado (fecha y hora)",
                "Creado de fecha y hora",
                "Creado",
                "Created",
                "Created at",
                "Fecha de creación",
                "Fecha y hora de creacion",
            ),
        ),
        ("Source", _pick_idx("Source", "Origen")),
        (
            "Actualizado (fecha y hora)",
            _pick_idx(
                "Actualizado (fecha y hora)",
                "fecha y hora de actualización",
                "Fecha y hora de actualización",
                "Actualizado",
                "Updated",
                "Updated at",
                "Fecha de actualización",
                "Fecha y hora de actualizacion",
            ),
        ),
        ("Archive", _pick_idx("Archive", "Archivado")),
        ("Shareable", _pick_idx("Shareable")),
        # Opcionales adicionales (si aparecen en tu tabla)
        ("Note", _pick_idx("Note", "Nota")),
        ("Sub-tipo de Ticket", _pick_idx("Sub-tipo de Ticket", "Sub-type", "Subtype")),
        ("Categoria del Cierre", _pick_idx("Categoria del Cierre", "Closure category", "Close category")),
        ("Promocion", _pick_idx("Promocion", "Promotion")),
    ]

    def _normalize_id_cliente(text: str) -> str:
        # En la UI viene como "R135921" o "G119230 (lead)".
        m = re.search(r"(\d+)", text)
        return m.group(1) if m else ""

    @dataclass(frozen=True)
    class _TableCfg:
        table_selector: str
        next_li_selector: str
        next_a_selector: str

    cfg = _TableCfg(
        table_selector=table_selector,
        next_li_selector="css=#admin_support_tickets_opened_list_next",
        next_a_selector="css=#admin_support_tickets_opened_list_next > a",
    )

    def read_page_rows() -> List[List[str]]:
        _wait_for_datatable_ready(scope, table_selector=cfg.table_selector, timeout_s=60.0)
        row_loc = table.locator("tbody tr")
        row_count = row_loc.count()

        page_rows: List[List[str]] = []
        for i in range(row_count):
            row = row_loc.nth(i)
            cells = row.locator("td")
            cell_count = cells.count()

            out_row: List[str] = []
            for col_name, zero_idx in col_map:
                if zero_idx is None or zero_idx < 0 or zero_idx >= cell_count:
                    out_row.append("")
                    continue
                txt = cells.nth(zero_idx).inner_text().strip()
                txt = " ".join(txt.split())
                if col_name == "ID Cliente":
                    txt = _normalize_id_cliente(txt)
                out_row.append(txt)

            if any(v for v in out_row):
                page_rows.append(out_row)
        return page_rows

    def page_marker() -> str:
        # Marca simple: primera fila completa. Útil para esperar cambio tras "Next".
        first_row = table.locator("tbody tr").first
        if first_row.count() == 0:
            return ""
        txt = first_row.inner_text().strip()
        return " ".join(txt.split())

    def is_next_enabled() -> bool:
        li = scope.locator(cfg.next_li_selector).first
        if li.count() == 0:
            return False
        cls = (li.get_attribute("class") or "").lower()
        return "disabled" not in cls and "paginate_button disabled" not in cls

    def click_next() -> None:
        scope.locator(cfg.next_a_selector).first.click()

    def wait_page_changed(old: str, timeout_s: float = 20.0) -> None:
        if not old:
            time.sleep(1.0)
            return
        start = time.monotonic()
        while True:
            if time.monotonic() - start > timeout_s:
                return
            try:
                new = page_marker()
            except Exception:
                new = ""
            if new and new != old:
                return
            time.sleep(0.25)

    wb = _open_or_create_workbook(output_xlsx)
    ws = _get_fresh_sheet(wb, sheet_name)
    ws.append([name for name, _ in col_map])

    # Recorre todas las páginas existentes
    while True:
        _wait_for_datatable_ready(scope, table_selector=cfg.table_selector, timeout_s=60.0)
        for r in read_page_rows():
            ws.append(r)

        if not is_next_enabled():
            break

        old = page_marker()
        click_next()
        wait_page_changed(old)

    wb.save(output_xlsx)


def extract_customers_to_excel(
    scope: LocatorScope,
    output_xlsx: str,
    sheet_name: str = "Datos Clientes",
    table_selector: str = "css=#customers_list_table",
    max_pages: int | None = None,
    max_rows: int | None = None,
) -> None:
    table = scope.locator(table_selector).first
    table.wait_for(state="visible")

    _wait_for_datatable_ready(scope, table_selector=table_selector, timeout_s=90.0)

    def _digits_longest(s: str) -> str:
        m = re.findall(r"\d+", s or "")
        return max(m, key=len) if m else ""

    def _normalize_spaces(s: str) -> str:
        return " ".join((s or "").strip().split())

    def _norm_header(s: str) -> str:
        s = _normalize_spaces(s).lower()
        s = re.sub(r"\s+", " ", s)
        return s

    def _build_header_index() -> dict[str, int]:
        """Mapea header visible -> índice 0-based del TD correspondiente."""
        try:
            ths = table.locator("thead th")
            texts = ths.all_inner_texts()
        except Exception:
            return {}

        header_to_idx: dict[str, int] = {}
        for idx, t in enumerate(texts):
            key = _norm_header(t)
            if key and key not in header_to_idx:
                header_to_idx[key] = idx
        return header_to_idx

    header_to_idx = _build_header_index()

    def _pick_col_index(fallback_one_based: int, *header_names: str) -> int:
        for name in header_names:
            key = _norm_header(name)
            if key in header_to_idx:
                return header_to_idx[key]
        return fallback_one_based - 1

    def _extract_customer_id_from_row(row) -> str:
        """Obtiene el ID real del cliente desde un enlace (si existe) o desde el texto."""
        # 1) Preferir hrefs que suelen contener el ID numérico
        try:
            links = row.locator('a[href*="customer"]').all()
        except Exception:
            links = []

        for a in links:
            try:
                href = a.get_attribute("href") or ""
                d = _digits_longest(href)
                if d:
                    return d
            except Exception:
                continue

        # 2) Fallback: usar la celda de la columna "ID" (si existe) y extraer dígitos
        try:
            id_idx = _pick_col_index(3, "ID")
            cells = row.locator("td")
            if 0 <= id_idx < cells.count():
                txt = _normalize_spaces(cells.nth(id_idx).inner_text())
                d = _digits_longest(txt)
                if d:
                    return d
        except Exception:
            pass

        return ""

    # Índices (1-based) como fallback; primero intentamos ubicar por el texto del header.
    # Se agregó la columna "Servicio usuario" (th:nth-child(9)).
    col_map = [
        ("Estado de Servicio", 2),
        ("ID", 3),
        ("Login del Portal", 4),
        ("Nombre Completo", 5),
        ("Número de Teléfono", 6),
        ("Tarifas de Internet", 7),
        ("Rangos IP", 8),
        ("Servicio usuario", 9),
        ("Socio", 10),
        ("Nacionalidad", 11),
        ("Estado", 12),
        ("Municipio", 13),
        ("Parroquia", 14),
        ("Residencia/Urbanización", 15),
    ]

    wb = _open_or_create_workbook(output_xlsx)
    ws = _get_fresh_sheet(wb, sheet_name)
    ws.append([name for name, _ in col_map])

    next_li_selector = "css=#customers_list_table_next"
    next_a_selector = "css=#customers_list_table_next > a"

    def read_page_rows() -> List[List[str]]:
        _wait_for_datatable_ready(scope, table_selector=table_selector, timeout_s=90.0)
        row_loc = table.locator("tbody tr")
        row_count = row_loc.count()

        page_rows: List[List[str]] = []
        for i in range(row_count):
            row = row_loc.nth(i)
            cells = row.locator("td")
            cell_count = cells.count()

            # ID: intentar leer el ID real (del href / texto) una sola vez por fila
            real_customer_id = _extract_customer_id_from_row(row)

            out_row: List[str] = []
            for col_name, one_based_idx in col_map:
                # Para ID, usamos el valor robusto (si existe)
                if col_name == "ID" and real_customer_id:
                    out_row.append(real_customer_id)
                    continue

                # Intentar ubicar la columna por header visible; si no, usar el índice fijo.
                zero_idx = _pick_col_index(
                    one_based_idx,
                    col_name,
                    # aliases frecuentes en Splynx/variantes
                    "estado de servicio" if col_name == "Estado de Servicio" else "",
                    "login" if col_name == "Login del Portal" else "",
                    "nombre" if col_name == "Nombre Completo" else "",
                    "numero" if col_name == "Número de Teléfono" else "",
                    "tarifas" if col_name == "Tarifas de Internet" else "",
                    "ip" if col_name == "Rangos IP" else "",
                    "servicio usario" if col_name == "Servicio usuario" else "",
                    "servicio usuario" if col_name == "Servicio usuario" else "",
                    "user service" if col_name == "Servicio usuario" else "",
                    "partner" if col_name == "Socio" else "",
                    "residencia" if col_name == "Residencia/Urbanización" else "",
                )
                if zero_idx < 0 or zero_idx >= cell_count:
                    out_row.append("")
                    continue
                txt = cells.nth(zero_idx).inner_text().strip()
                txt = " ".join(txt.split())
                out_row.append(txt)

            if any(v for v in out_row):
                page_rows.append(out_row)

        return page_rows

    def page_marker() -> str:
        first_row = table.locator("tbody tr").first
        if first_row.count() == 0:
            return ""
        txt = first_row.inner_text().strip()
        return " ".join(txt.split())

    def is_next_enabled() -> bool:
        li = scope.locator(next_li_selector).first
        if li.count() == 0:
            return False
        cls = (li.get_attribute("class") or "").lower()
        return "disabled" not in cls and "paginate_button disabled" not in cls

    def click_next() -> None:
        scope.locator(next_a_selector).first.click()

    def wait_page_changed(old: str, timeout_s: float = 25.0) -> None:
        if not old:
            time.sleep(1.0)
            return
        start = time.monotonic()
        while True:
            if time.monotonic() - start > timeout_s:
                return
            try:
                new = page_marker()
            except Exception:
                new = ""
            if new and new != old:
                return
            time.sleep(0.25)

    pages_done = 0
    rows_written = 0

    # Recorre páginas (o limita por max_pages/max_rows si se indican)
    while True:
        page_rows = read_page_rows()
        for r in page_rows:
            ws.append(r)
            rows_written += 1
            if isinstance(max_rows, int) and max_rows > 0 and rows_written >= max_rows:
                wb.save(output_xlsx)
                return

        pages_done += 1
        if isinstance(max_pages, int) and max_pages > 0 and pages_done >= max_pages:
            break

        if not is_next_enabled():
            break

        old = page_marker()
        click_next()
        wait_page_changed(old)

    wb.save(output_xlsx)


def _wait_for_datatable_ready(scope: LocatorScope, table_selector: str, timeout_s: float) -> None:
    table = scope.locator(table_selector).first
    table.wait_for(state="visible")

    start = time.monotonic()
    while True:
        if time.monotonic() - start > timeout_s:
            return

        # Si existe overlay de "processing", esperar a que desaparezca.
        try:
            proc = scope.locator("css=div.dataTables_processing").first
            if proc.count() > 0 and proc.is_visible():
                time.sleep(0.25)
                continue
        except Exception:
            pass

        try:
            # Caso normal: filas presentes
            if table.locator("tbody tr").count() > 0:
                # Caso vacío (DataTables): una sola fila con td.dataTables_empty
                if table.locator("tbody tr td.dataTables_empty").count() > 0:
                    return
                return
        except Exception:
            pass

        time.sleep(0.25)


def _open_or_create_workbook(path: str) -> Workbook:
    p = Path(path)
    if not p.exists():
        return Workbook()

    try:
        return load_workbook(path)
    except PermissionError as exc:
        raise PermissionError(
            f"No se puede abrir '{path}'. Probablemente está abierto en Excel o bloqueado. "
            "Ciérralo y vuelve a intentar."
        ) from exc
    except Exception as exc:
        # Importante: NO crear un archivo nuevo si el existente no se pudo leer,
        # porque eso termina sobrescribiendo el Excel y perdiendo pestañas.
        raise RuntimeError(
            f"No se pudo leer el Excel existente '{path}'. No se modificó el archivo. Detalle: {exc}"
        ) from exc


def _get_fresh_sheet(wb: Workbook, sheet_name: str):
    if sheet_name in wb.sheetnames:
        ws_old = wb[sheet_name]
        wb.remove(ws_old)

    # Si el workbook recién creado solo tiene la hoja por defecto y está vacía, la reutilizamos.
    if wb.sheetnames == ["Sheet"] and wb["Sheet"].max_row == 1 and wb["Sheet"].max_column == 1:
        ws = wb["Sheet"]
        ws.title = sheet_name
        return ws

    return wb.create_sheet(title=sheet_name)
